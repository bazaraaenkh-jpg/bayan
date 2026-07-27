"""Үе шат 3-ын Тестүүд: FIFO Бараа материал, Агуулах, Setup Wizard болон Эхний үлдэгдэл импортлох.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
import pytest
from openpyxl import Workbook

from sqlalchemy import select
from bayan import inventory, ledger
from bayan.coa_seed import setup_company, seed_company
from bayan.models import Account, Company, NormalSide
from bayan.opening_balances import import_opening_balances


def test_fifo_costing(session):
    # 1. FIFO тохиргоотой компани үүсгэх
    c = setup_company(session, "FIFO Тესт", "retail", is_vat_payer=False, inventory_method="fifo")
    
    # 2. Бараа үүсгэх
    item = inventory.Item(company_id=c.id, code="B1", name="Бараа 1", unit="ш", gl_account="2101")
    session.add(item)
    session.flush()

    # 3. Багц 1: 10 ширхэг, нэгж нь 1,000₮ (нийт 10,000₮ = 1,000,000 мөнгө)
    inventory.receive(session, c.id, item, date(2026, 3, 1), qty=10, total_cost_minor=1_000_000)
    
    # 4. Багц 2: 10 ширхэг, нэгж нь 1,500₮ (нийт 15,000₮ = 1,500,000 мөнгө)
    inventory.receive(session, c.id, item, date(2026, 3, 5), qty=10, total_cost_minor=1_500_000)

    # Баланс шалгах: нийт 20 ширхэг, өртөг 25,000₮
    assert item.qty == 20
    assert item.total_cost_minor == 2_500_000

    # 5. 12 ширхэг зарлагадах (FIFO-оор: 10ш * 1000₮ + 2ш * 1500₮ = 13,000₮ = 1,300,000 мөнгө)
    move = inventory.issue(session, c.id, item, date(2026, 3, 10), qty=12)

    assert move.cost_minor == 1_300_000
    assert item.qty == 8
    assert item.total_cost_minor == 1_200_000  # Үлдсэн 8ш нь дандаа 1500₮-ийн багцаас (8 * 1500 = 12000₮)


def test_warehouse_transfer(session):
    c = setup_company(session, "Warehouse Тესт", "retail", is_vat_payer=False, inventory_method="fifo")
    
    wh1 = inventory.Warehouse(company_id=c.id, code="WH01", name="Төв агуулах")
    wh2 = inventory.Warehouse(company_id=c.id, code="WH02", name="Салбар агуулах")
    session.add_all([wh1, wh2])
    
    item = inventory.Item(company_id=c.id, code="B2", name="Бараа 2", unit="ш", gl_account="2101")
    session.add(item)
    session.flush()

    # WH1 рүү 10ш бараа орлого авах (нэгж нь 2,000₮, нийт 20,000₮)
    inventory.receive(session, c.id, item, date(2026, 3, 1), qty=10, total_cost_minor=2_000_000, warehouse_id=wh1.id)

    # WH1-ээс WH2 рүү 4ш бараа шилжүүлэх
    moves = inventory.transfer(session, c.id, item, from_warehouse_id=wh1.id, to_warehouse_id=wh2.id, qty=4, move_date=date(2026, 3, 5))

    assert len(moves) == 2
    assert moves[0].kind == inventory.MoveKind.issue and moves[0].warehouse_id == wh1.id
    assert moves[1].kind == inventory.MoveKind.receipt and moves[1].warehouse_id == wh2.id
    assert moves[0].cost_minor == 800_000  # 4 * 2000₮ = 8000₮ = 800,000 мөнгө

    # Агуулахын тайлан шалгах
    r_wh1 = inventory.warehouse_stock_report(session, c.id, warehouse_id=wh1.id)
    assert r_wh1[0]["qty"] == 6
    assert r_wh1[0]["total_cost_minor"] == 1_200_000

    r_wh2 = inventory.warehouse_stock_report(session, c.id, warehouse_id=wh2.id)
    assert r_wh2[0]["qty"] == 4
    assert r_wh2[0]["total_cost_minor"] == 800_000


def test_company_setup_wizard(session):
    # Үйлдвэрлэл + НӨАТ төлөгч
    c1 = setup_company(session, "Үйлдвэр Компани", "manufacturing", is_vat_payer=True)
    codes1 = {a.code for a in session.scalars(select(Account).where(Account.company_id == c1.id)).all()}
    
    assert "2145" in codes1  # Дуусаагүй үйлдвэрлэл
    assert "2151" in codes1  # Бэлэн бүтээгдэхүүн
    assert "1203" in codes1  # НӨАТ-ын авлага
    assert "3105" in codes1  # НӨАТ-ын өглөг

    # Үйлчилгээ + НӨАТ төлөгч биш
    c2 = setup_company(session, "Үйлчилгээ Компани", "service", is_vat_payer=False)
    codes2 = {a.code for a in session.scalars(select(Account).where(Account.company_id == c2.id)).all()}
    
    assert "2101" not in codes2 # Бараа материал байхгүй
    assert "2145" not in codes2
    assert "1203" not in codes2 # НӨАТ-ын авлага байхгүй
    assert "6102" in codes2     # Үйлчилгээний өртөг данс үүссэн


def test_import_opening_balances(session, tmp_path):
    c = setup_company(session, "Баланс Компани", "retail", is_vat_payer=False)
    
    # Синтетик Excel файл үүсгэх
    f = tmp_path / "opening_balances.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Дансны код", "Дебит үлдэгдэл", "Кредит үлдэгдэл"])
    ws.append(["1001", "1,500,000.00", None]) # Касс
    ws.append(["4101", None, "1,500,000.00"]) # Өмч
    wb.save(f)

    # Импорт хийх (Баланс тэнцүү: 1,500,000₮)
    res = import_opening_balances(session, c.id, f, date(2025, 12, 31))
    
    assert res["entry_id"] is not None
    assert res["lines_count"] == 2
    assert res["total_minor"] == 150_000_000  # 1.5 сая төгрөг

    # Баланс зөрүүтэй Excel файл үүсгэж алдаа шалгах
    f_err = tmp_path / "opening_error.xlsx"
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(["Данс", "Дебит", "Кредит"])
    ws2.append(["1001", "1,500,000.00", None])
    ws2.append(["4101", None, "1,200,000.00"]) # 300,000₮ зөрүүтэй
    wb2.save(f_err)

    with pytest.raises(ValueError, match="Баланс зөрүүтэй байна"):
        import_opening_balances(session, c.id, f_err, date(2025, 12, 31))


def test_period_lock_ledger(session):
    c = setup_company(session, "Lock Company", "retail", is_vat_payer=False)
    
    # 1. Сар түгжих
    ledger.lock_period(session, c.id, 2026, 3, actor_id="user-1")
    
    # 2. Түгжсэн сард бичилт хийхэд алдаа заана
    lines = [
        ledger.LineInput(account_code="1001", debit_minor=1000, description="Туршилт"),
        ledger.LineInput(account_code="4101", credit_minor=1000, description="Туршилт"),
    ]
    with pytest.raises(ledger.PeriodLockedError, match="сар түгжээтэй байна"):
        ledger.post_entry(session, c.id, date(2026, 3, 15), lines, actor_id="user-1")


def test_company_guard_subscription(session):
    from fastapi import HTTPException
    from bayan.api import company_guard
    from bayan.auth import add_membership
    from bayan.models import Subscription
    from sqlalchemy import delete
    
    c = setup_company(session, "Sub Guard Company", "retail", is_vat_payer=False)
    add_membership(session, "user-1", c.id, "owner")
    
    # Идэвхтэй багцыг устгана (setup_company дээр үүсгэсэн TRIAL-ийг устгах)
    session.execute(delete(Subscription).where(Subscription.company_id == c.id))
    session.flush()
    
    guard = company_guard("post")
    
    class MockRequest:
        class url:
            path = "/api/companies/some-id/receive"
            
    mock_request = MockRequest()
    
    # 1. Багцгүй тул 403 алдаа өгнө
    with pytest.raises(HTTPException) as exc:
        guard(c.id, mock_request, {"uid": "user-1"}, session)
    assert exc.value.status_code == 403
    assert "багцын хугацаа дууссан" in exc.value.detail
    
    # 2. Багц нэмэх
    from datetime import datetime, timedelta
    sub = Subscription(
        company_id=c.id,
        plan="PREMIUM",
        starts_at=datetime.utcnow() - timedelta(days=1),
        ends_at=datetime.utcnow() + timedelta(days=30),
        status="ACTIVE"
    )
    session.add(sub)
    session.flush()
    
    # 3. Багцтай болсон тул амжилттай нэвтрүүлнэ
    res = guard(c.id, mock_request, {"uid": "user-1"}, session)
    assert res["company_id"] == c.id


def test_llm_descriptor_generator(session, tmp_path):
    from bayan import llm_descriptor
    import yaml
    
    # 1. Excel хуулга үүсгэх
    f = tmp_path / "khan_bank_statement.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Гүйлгээний огноо", "Харилцагч", "Зарлага", "Орлого", "Гүйлгээний утга"])
    ws.append(["2026-03-01", "Алимаа", "1,500.00", None, "Шилжүүлэг"])
    ws.append(["2026-03-02", "Бат", None, "2,000.00", "Орлого орлоо"])
    wb.save(f)
    
    # 2. YAML үүсгэх
    yaml_text = llm_descriptor.generate_yaml_descriptor(f, "khan_bank_statement.xlsx", "khan")
    
    # 3. Баталгаажуулах
    data = yaml.safe_load(yaml_text)
    assert data["bank"] == "khan"
    assert data["file_type"] == "excel"
    assert "date" in data["table"]["columns"]
    assert "description" in data["table"]["columns"]


def test_ebarimt_bank_txn_matching(session):
    from datetime import datetime
    from bayan.classify import classify_batch
    from bayan.partners import Invoice, InvoiceKind, Counterparty
    from bayan.models import BankTxn, Direction, ExtractionPath, Statement, StatementStatus
    
    # 1. Сэтэп компани ба харилцагч
    c = setup_company(session, "E-Barimt Matching Co", "retail", is_vat_payer=True)
    cp = Counterparty(company_id=c.id, name="Supplier Co", reg_no="9999999")
    session.add(cp)
    session.flush()
    
    # 2. И-Баримтаас орсон өглөг (Purchase Invoice) үүсгэх
    inv = Invoice(
        company_id=c.id,
        counterparty_id=cp.id,
        kind=InvoiceKind.purchase,
        number="INV-2026-0001",
        issue_date=date(2026, 3, 10),
        due_date=date(2026, 3, 10),
        net_minor=90_000,   # 900₮ net
        vat_minor=10_000,   # 100₮ vat (Нийт 1000₮ = 100,000 minor)
        paid_minor=0,
    )
    session.add(inv)
    session.flush()
    
    # 3. Банкны гүйлгээ үүсгэх (Outgoing/Debit 100,000 minor)
    stmt = Statement(
        company_id=c.id,
        file_name="stmt.xlsx",
        file_sha256="fake_sha",
        status=StatementStatus.uploaded
    )
    session.add(stmt)
    session.flush()
    
    txn = BankTxn(
        statement_id=stmt.id,
        company_id=c.id,
        bank_account_key="5011",
        seq_no=1,
        posted_at=datetime(2026, 3, 11, 10, 0),
        direction=Direction.debit,
        amount_minor=100_000,   # 1000₮
        description_raw="Supplier Co payment",
        description_norm="Supplier Co payment",
        canonical_hash="hash1",
        extraction_path=ExtractionPath.excel,
    )
    session.add(txn)
    session.flush()
    
    # 4. Ангилах (classify_batch ажиллуулна)
    suggestions = classify_batch(session, c.id, [txn], use_ai=False)
    
    assert len(suggestions) == 1
    s = suggestions[0]
    assert s.account_code == "2101" # Accounts Payable suggested!
    assert s.vat_flag is True       # Matches ebarimt vat
    assert s.confidence == 0.98     # Matches ebarimt match confidence
    assert "И-Баримтын дүнтэй таарлаа" in s.rationale


def test_new_roles_auth(session):
    from bayan import auth
    
    # 1. Chief Accountant can do everything (admin, post, approve, read)
    auth.require("chief_accountant", "post")
    auth.require("chief_accountant", "approve")
    auth.require("chief_accountant", "admin")
    auth.require("chief_accountant", "read")
    
    # 2. Cashier can post and read, but not admin/approve
    auth.require("cashier", "post")
    auth.require("cashier", "read")
    with pytest.raises(auth.AuthError):
        auth.require("cashier", "approve")
    with pytest.raises(auth.AuthError):
        auth.require("cashier", "admin")
        
    # 3. Warehouse Manager can post and read, but not admin/approve
    auth.require("warehouse_manager", "post")
    auth.require("warehouse_manager", "read")
    with pytest.raises(auth.AuthError):
        auth.require("warehouse_manager", "approve")
    with pytest.raises(auth.AuthError):
        auth.require("warehouse_manager", "admin")


def test_user_two_factor_auth(session):
    from bayan import auth
    
    # 1. Хэрэглэгч үүсгэх
    email = "2fa_test@bayan.ai"
    u = auth.create_user(session, email, "2FA User", "password123")
    
    assert u.totp_enabled is False
    assert u.totp_secret is None
    
    # 2. Нэвтрэх: 2FA идэвхгүй үед шууд токен өгнө
    token = auth.login(session, email, "password123")
    assert token is not None
    
    # 3. 2FA нууц үг үүсгэж тохируулах
    secret = auth.generate_totp_secret()
    u.totp_secret = secret
    session.flush()
    
    # Мок TOTP код бодох
    import time
    import hmac
    import hashlib
    import struct
    import base64
    
    key = base64.b32decode(secret.upper().encode())
    t = int(time.time() // 30)
    msg = struct.pack(">Q", t)
    h = hmac.new(key, msg, hashlib.sha1).digest()
    o = h[19] & 15
    token_val = (struct.unpack(">I", h[o:o+4])[0] & 0x7fffffff) % 1000000
    code = f"{token_val:06d}"
    
    # Код зөв болохыг шалгана
    assert auth.verify_totp(secret, code) is True
    
    # 2FA-г идэвхжүүлэх
    u.totp_enabled = True
    session.flush()
    
    # 4. 2FA идэвхжсэн тул ерөнхий login алдаа заана (2FA_REQUIRED)
    with pytest.raises(auth.AuthError, match="2FA_REQUIRED"):
        auth.login(session, email, "password123")


def test_subscription_limits(session):
    from fastapi import HTTPException
    from bayan.api import create_warehouse, WarehouseIn
    from bayan.models import Subscription
    from datetime import datetime, timedelta
    from sqlalchemy import delete
    
    c = setup_company(session, "Limit Test Co", "retail", is_vat_payer=False)
    
    # 1. TRIAL limit test (1 warehouse max)
    session.execute(delete(Subscription).where(Subscription.company_id == c.id))
    trial_sub = Subscription(
        company_id=c.id,
        plan="TRIAL",
        starts_at=datetime.utcnow() - timedelta(days=1),
        ends_at=datetime.utcnow() + timedelta(days=30),
        status="ACTIVE"
    )
    session.add(trial_sub)
    session.flush()
    
    # Create first warehouse: should succeed
    res = create_warehouse(c.id, WarehouseIn(code="WH1", name="Warehouse 1"), {"uid": "user-1"}, session)
    assert res["code"] == "WH1"
    
    # Create second warehouse: should fail on TRIAL
    with pytest.raises(HTTPException) as exc:
        create_warehouse(c.id, WarehouseIn(code="WH2", name="Warehouse 2"), {"uid": "user-1"}, session)
    assert exc.value.status_code == 400
    assert "багцын хязгаар хэтэрсэн" in exc.value.detail


def test_financial_reports_vat_posting_reversal(session):
    from bayan.partners import post_invoice, InvoiceKind, Counterparty
    from bayan.vat import tt03a
    from bayan.models import EntryStatus, JournalEntry
    
    # 1. Setup company & partner
    c = setup_company(session, "Financial Audit Co", "retail", is_vat_payer=True)
    cp = Counterparty(company_id=c.id, name="Test Partner", reg_no="1234567")
    session.add(cp)
    session.flush()
    
    # 2. Post Sales Invoice (Bortluulalt) - 100,000 minor (1,000₮), with 10% VAT
    sales_inv = post_invoice(
        session, c.id, cp.id,
        kind=InvoiceKind.sales,
        number="SALE-001",
        issue_date=date(2026, 3, 10),
        due_date=date(2026, 3, 10),
        net_minor=100000,
        with_vat=True
    )
    
    # 3. Post Purchase Invoice (Huldaldan avalt) - 50,000 minor (500₮), with 10% VAT
    purch_inv = post_invoice(
        session, c.id, cp.id,
        kind=InvoiceKind.purchase,
        number="PURCH-001",
        issue_date=date(2026, 3, 15),
        due_date=date(2026, 3, 15),
        net_minor=50000,
        with_vat=True,
        expense_account="7101"
    )
    
    # 4. Generate and verify VAT TT-03A report
    vat_rep = tt03a(session, c.id, 2026, 3)
    rows = vat_rep["rows"]
    
    assert rows["1_niit_borluulalt"] == 100000
    assert rows["26_nogduulsan_tatvar"] == 10000
    assert rows["32_niit_hudaldan_avalt"] == 55000  # net + vat = 50,000 + 5,000 = 55,000
    assert rows["42_tolson_noat"] == 5000
    assert rows["64_etssiin_tolboh"] == 5000        # 10,000 - 5,000 = 5,000₮ payable
    
    # 5. Reverse the sales invoice:
    orig_entry = session.get(JournalEntry, sales_inv.journal_entry_id)
    assert orig_entry.status == EntryStatus.posted
    
    # Perform reversal (cancellation)
    rev_entry = ledger.reverse_entry(session, sales_inv.journal_entry_id, actor_id="user-1")
    
    # Check that original entry is now reversed and the new entry is posted
    assert orig_entry.status == EntryStatus.reversed
    assert rev_entry.status == EntryStatus.posted
    assert rev_entry.reversal_of == orig_entry.id


def test_company_settings_and_vat_activation(session):
    from bayan.api import update_company_settings, CompanySettingsIn
    from bayan.models import Account
    from sqlalchemy import select
    
    # Create non-VAT paying company
    c = setup_company(session, "Non VAT Co", "retail", is_vat_payer=False)
    
    # Check that VAT accounts (3105 / 1203) do not exist
    vat_3105 = session.scalar(select(Account).where(Account.company_id == c.id, Account.code == "3105"))
    vat_1203 = session.scalar(select(Account).where(Account.company_id == c.id, Account.code == "1203"))
    assert vat_3105 is None
    assert vat_1203 is None
    
    # Call settings update endpoint to change vat_payer to True
    res = update_company_settings(c.id, CompanySettingsIn(vat_payer=True), {"uid": "user-1"}, session)
    assert res["vat_payer"] is True
    
    # Check that VAT accounts have been created and initialized
    vat_3105 = session.scalar(select(Account).where(Account.company_id == c.id, Account.code == "3105"))
    vat_1203 = session.scalar(select(Account).where(Account.company_id == c.id, Account.code == "1203"))
    assert vat_3105 is not None
    assert vat_1203 is not None
    assert vat_3105.is_postable is True
    assert vat_1203.is_postable is True


def test_financial_ratios_calculation(session):
    from bayan.ratios import calculate_financial_ratios
    from bayan.partners import post_invoice, InvoiceKind, Counterparty
    
    # 1. Setup company & partner
    c = setup_company(session, "Ratios Co", "retail", is_vat_payer=False)
    cp = Counterparty(company_id=c.id, name="Ratio Partner", reg_no="7654321")
    session.add(cp)
    session.flush()
    
    # Post some transactions to build balances
    # Post sales: 1,000₮
    post_invoice(session, c.id, cp.id, InvoiceKind.sales, "S-1", date(2026, 3, 1), date(2026, 3, 1), 100000)
    # Post expense: 400₮
    post_invoice(session, c.id, cp.id, InvoiceKind.purchase, "P-1", date(2026, 3, 5), date(2026, 3, 5), 40000, expense_account="7101")
    
    # 2. Calculate financial ratios
    ratios = calculate_financial_ratios(session, c.id)
    
    # 3. Verify ratios structure and calculations
    assert "working_capital" in ratios
    assert "current_ratio" in ratios
    assert "quick_ratio" in ratios
    assert "gross_margin" in ratios
    assert "net_margin" in ratios
    assert "roa" in ratios
    assert "roe" in ratios
    assert "debt_ratio" in ratios
    
    # Verify values (Net Profit = 600₮, Revenue = 1,000₮)
    assert ratios["net_margin"]["value"] == 60.0  # 60% net profit margin
    assert ratios["working_capital"]["value"] >= 0
    assert ratios["working_capital"]["status"] in ("Сайн", "Анхаарах")


def test_timeline_financial_summary(session):
    from bayan.ratios import get_period_financial_summary
    from bayan.partners import post_invoice, InvoiceKind, Counterparty
    
    # 1. Setup company & partner
    c = setup_company(session, "Timeline Co", "retail", is_vat_payer=False)
    cp = Counterparty(company_id=c.id, name="Timeline Partner", reg_no="9998887")
    session.add(cp)
    session.flush()
    
    # Post some transactions in different dates
    # Sales in March: 1,000₮
    post_invoice(session, c.id, cp.id, InvoiceKind.sales, "S-1", date(2026, 3, 10), date(2026, 3, 10), 100000)
    # Expense in April: 300₮
    post_invoice(session, c.id, cp.id, InvoiceKind.purchase, "P-1", date(2026, 4, 15), date(2026, 4, 15), 30000, expense_account="7101")
    
    # Query summary specifically for March (date_from=2026-03-01, date_to=2026-03-31)
    summary_march = get_period_financial_summary(session, c.id, date(2026, 3, 1), date(2026, 3, 31))
    assert summary_march["revenue"] == 1000.0
    assert summary_march["expenses"] == 0.0
    assert summary_march["net_income"] == 1000.0
    
    # Query summary specifically for April (date_from=2026-04-01, date_to=2026-04-30)
    summary_april = get_period_financial_summary(session, c.id, date(2026, 4, 1), date(2026, 4, 30))
    assert summary_april["revenue"] == 0.0
    assert summary_april["expenses"] == 300.0
    assert summary_april["net_income"] == -300.0


def test_unified_dashboard_summary(session):
    from bayan.ratios import get_unified_dashboard_summary
    from bayan.partners import post_invoice, InvoiceKind, Counterparty
    
    c = setup_company(session, "Unified Co", "retail", is_vat_payer=False)
    cp = Counterparty(company_id=c.id, name="Unified Partner", reg_no="1112223")
    session.add(cp)
    session.flush()
    
    # Post transactions
    post_invoice(session, c.id, cp.id, InvoiceKind.sales, "S-1", date(2026, 3, 10), date(2026, 3, 10), 200000)
    post_invoice(session, c.id, cp.id, InvoiceKind.purchase, "P-1", date(2026, 3, 15), date(2026, 3, 15), 50000, expense_account="7101")
    
    # Query unified dashboard
    dash = get_unified_dashboard_summary(session, c.id, date(2026, 3, 1), date(2026, 3, 31))
    
    assert "ratios" in dash
    assert "summary" in dash
    assert dash["summary"]["revenue"] == 2000.0
    assert dash["summary"]["net_income"] == 1500.0
    assert dash["ratios"]["net_margin"]["value"] == 75.0


def test_dashboard_alerts_negative_balances(session):
    from bayan.ratios import get_unified_dashboard_summary
    from bayan.partners import post_invoice, InvoiceKind, Counterparty
    
    c = setup_company(session, "Alerts Co", "retail", is_vat_payer=False)
    cp = Counterparty(company_id=c.id, name="Alerts Partner", reg_no="5556667")
    session.add(cp)
    session.flush()
    
    # Post purchase (expense) of 500₮ without any sales or capital.
    # This results in: Accounts Payable Credit 500₮, Expense Debit 500₮.
    # And we pay it from Cash (which will go negative Dt 1001 cash = -500₮)
    # Let's post a direct payment to cp: Dt 3101 (Payables) 500₮, Kt 1001 (Cash) 500₮.
    from bayan.ledger import post_entry, LineInput
    post_entry(session, c.id, date(2026, 3, 5), [
        LineInput(account_code="3101", debit_minor=50000),
        LineInput(account_code="1001", credit_minor=50000)
    ])
    
    dash = get_unified_dashboard_summary(session, c.id, date(2026, 3, 1), date(2026, 3, 31))
    assert "alerts" in dash
    alerts_types = [a["type"] for a in dash["alerts"]]
    assert "CASH_NEGATIVE" in alerts_types


def test_period_lock_checklist_validation(session):
    from bayan.ledger import lock_period, LedgerError
    from bayan.models import ClassificationSuggestion, BankTxn, Statement, StatementStatus, ExtractionPath, Direction
    from datetime import datetime
    
    c = setup_company(session, "Lock Co", "retail", is_vat_payer=False)
    
    # 1. Initially, it should lock successfully because balance is 0 = 0
    lock_period(session, c.id, 2026, 3)
    
    # 2. Let's create another company and add a pending suggestion to test validation failure
    c2 = setup_company(session, "Lock Co 2", "retail", is_vat_payer=False)
    
    stmt = Statement(
        company_id=c2.id,
        file_name="stmt.xlsx",
        file_sha256="abc12345",
        period_from=date(2026, 3, 1),
        period_to=date(2026, 3, 31),
        status=StatementStatus.uploaded
    )
    session.add(stmt)
    session.flush()
    
    txn = BankTxn(
        statement_id=stmt.id,
        company_id=c2.id,
        bank_account_key="khan_123",
        seq_no=1,
        posted_at=datetime(2026, 3, 10),
        direction=Direction.credit,
        amount_minor=10000,
        description_raw="test",
        description_norm="test",
        canonical_hash="hash1",
        extraction_path=ExtractionPath.excel
    )
    session.add(txn)
    session.flush()
    
    sugg = ClassificationSuggestion(
        bank_txn_id=txn.id,
        company_id=c2.id,
        account_code="7101",
        confidence=1.0,
        source="rule",
        status="pending"
    )
    session.add(sugg)
    session.flush()
    
    # Try to lock 2026-03 for c2; should raise LedgerError because of pending suggestion
    import pytest
    with pytest.raises(LedgerError) as excinfo:
        lock_period(session, c2.id, 2026, 3)
    assert "батлагдаагүй" in str(excinfo.value)


def test_legal_compliance_features(session):
    from bayan.ledger import lock_period
    from bayan.models import PeriodLock
    from bayan.auth import User
    from sqlalchemy import select
    
    # 1. Setup company and CPA user
    c = setup_company(session, "Compliance Co", "retail", is_vat_payer=False)
    
    # 2. Lock period with CPA license
    lock_period(session, c.id, 2026, 3, actor_id="user-123", cpa_license_no="CPA-778899")
    
    # 3. Verify license number is saved
    lock_rec = session.scalar(
        select(PeriodLock).where(PeriodLock.company_id == c.id, PeriodLock.year == 2026, PeriodLock.month == 3)
    )
    assert lock_rec is not None
    assert lock_rec.cpa_license_no == "CPA-778899"
    
    # 4. Verify export logic returns dictionary of data
    from bayan.api import get_company_legal_archive
    # mock company_guard admin dependency check and context
    # we can call the service directly with mock session
    archive = get_company_legal_archive(c.id, ctx={"uid": "user-123", "role": "chief_accountant"}, db=session)
    assert "company" in archive
    assert "period_locks" in archive
    assert "accounts" in archive
    assert "journal_entries" in archive
    assert "bank_statements" in archive
    assert archive["company"]["name"] == "Compliance Co"
    assert archive["period_locks"][0]["cpa_license_no"] == "CPA-778899"


def test_owner_cash_runway_and_burn_rate(session):
    from bayan.ratios import get_unified_dashboard_summary
    from bayan.partners import post_invoice, InvoiceKind, Counterparty
    from bayan.ledger import post_entry, LineInput
    
    c = setup_company(session, "Owner Co", "retail", is_vat_payer=False)
    cp = Counterparty(company_id=c.id, name="Owner Partner", reg_no="3334445")
    session.add(cp)
    session.flush()
    
    # 1. Invest Capital: Cash 10000₮, Equity 10000₮
    post_entry(session, c.id, date(2026, 3, 1), [
        LineInput("1001", debit_minor=1000000),
        LineInput("4101", credit_minor=1000000)
    ])
    # 2. Expense: 2000₮
    post_invoice(session, c.id, cp.id, InvoiceKind.purchase, "P-1", date(2026, 3, 5), date(2026, 3, 5), 200000, expense_account="7101")
    
    # Get dashboard
    dash = get_unified_dashboard_summary(session, c.id, date(2026, 3, 1), date(2026, 3, 31))
    
    assert "summary" in dash
    assert "burn_rate" in dash["summary"]
    assert "cash_runway_months" in dash["summary"]
    assert dash["summary"]["burn_rate"] > 0
    assert dash["summary"]["cash_runway_months"] != "Хязгааргүй"


def test_owner_expense_variance_alerts(session):
    from bayan.ratios import get_unified_dashboard_summary
    from bayan.partners import post_invoice, InvoiceKind, Counterparty
    
    c = setup_company(session, "Variance Co", "retail", is_vat_payer=False)
    cp = Counterparty(company_id=c.id, name="Variance Partner", reg_no="2223334")
    session.add(cp)
    session.flush()
    
    # Post March expense: 100₮
    post_invoice(session, c.id, cp.id, InvoiceKind.purchase, "P-1", date(2026, 3, 10), date(2026, 3, 10), 10000, expense_account="7101")
    # Post April expense: 200₮ (100% MoM increase!)
    post_invoice(session, c.id, cp.id, InvoiceKind.purchase, "P-2", date(2026, 4, 15), date(2026, 4, 15), 20000, expense_account="7101")
    
    # Query April
    dash = get_unified_dashboard_summary(session, c.id, date(2026, 4, 1), date(2026, 4, 30))
    alerts_types = [a["type"] for a in dash["alerts"]]
    assert "EXPENSE_VARIANCE" in alerts_types


def test_ceo_invoice_approval_gate(session):
    from bayan.api import create_invoice, InvoiceIn
    from fastapi import HTTPException
    import pytest
    
    c = setup_company(session, "CEO Co", "retail", is_vat_payer=False)
    from bayan.partners import Counterparty
    cp = Counterparty(company_id=c.id, name="CEO Partner", reg_no="1234567")
    session.add(cp)
    session.flush()
    
    # 6,000,000 MNT = 600,000,000 minor units
    body = InvoiceIn(
        counterparty_id=cp.id,
        kind="purchase",
        number="P-99",
        issue_date=date(2026, 3, 1),
        due_date=date(2026, 3, 31),
        net_minor=600000000
    )
    
    # Try creating as accountant (not owner/chief_accountant); should raise HTTP 400
    with pytest.raises(HTTPException) as excinfo:
        create_invoice(c.id, body, ctx={"uid": "user-1", "role": "accountant"}, db=session)
    assert excinfo.value.status_code == 400
    assert "Удирдлагын зөвшөөрөл" in excinfo.value.detail
    
    # Create as owner; should succeed
    res = create_invoice(c.id, body, ctx={"uid": "user-1", "role": "owner"}, db=session)
    assert "id" in res


def test_audit_segregation_of_duties_invoice(session):
    from bayan.api import create_invoice, InvoiceIn
    from bayan.partners import Counterparty
    from fastapi import HTTPException
    import pytest
    
    c = setup_company(session, "SoD Co", "retail", is_vat_payer=False)
    # Counterparty created by user-123
    cp = Counterparty(company_id=c.id, name="SoD Supplier", reg_no="9998887", created_by="user-123")
    session.add(cp)
    session.flush()
    
    body = InvoiceIn(
        counterparty_id=cp.id,
        kind="purchase",
        number="P-101",
        issue_date=date(2026, 3, 1),
        due_date=date(2026, 3, 31),
        net_minor=10000
    )
    
    # Try posting invoice as user-123 (SoD Conflict!); should raise HTTP 400
    with pytest.raises(HTTPException) as excinfo:
        create_invoice(c.id, body, ctx={"uid": "user-123", "role": "accountant"}, db=session)
    assert excinfo.value.status_code == 400
    assert "Үүргийн тусгаарлалт" in excinfo.value.detail
    
    # Try as user-999; should succeed
    res = create_invoice(c.id, body, ctx={"uid": "user-999", "role": "accountant"}, db=session)
    assert "id" in res


def test_audit_sequence_gap_detection(session):
    from bayan.ledger import post_entry, LineInput, lock_period, LedgerError
    from bayan.models import JournalEntry
    from sqlalchemy import select
    import pytest
    
    c = setup_company(session, "Gap Co", "retail", is_vat_payer=False)
    
    # Post first entry
    post_entry(session, c.id, date(2026, 3, 1), [
        LineInput("1001", debit_minor=1000),
        LineInput("4101", credit_minor=1000)
    ])
    # Post second entry
    e2 = post_entry(session, c.id, date(2026, 3, 2), [
        LineInput("1001", debit_minor=2000),
        LineInput("4101", credit_minor=2000)
    ])
    
    # Manually modify entry_no of e2 to 3, leaving a gap (1, 3)
    e2.entry_no = 3
    session.flush()
    
    # Attempt to lock period; should raise LedgerError with Sequence Gap details
    with pytest.raises(LedgerError) as excinfo:
        lock_period(session, c.id, 2026, 3)
    assert "Sequence Gap" in str(excinfo.value)


def test_audit_is_system_generated_flag(session):
    from bayan.ledger import post_entry, LineInput
    from bayan.models import JournalEntry
    
    c = setup_company(session, "SysGen Co", "retail", is_vat_payer=False)
    
    e = post_entry(session, c.id, date(2026, 3, 1), [
        LineInput("1001", debit_minor=1000),
        LineInput("4101", credit_minor=1000)
    ], is_system_generated=True)
    
    assert e.is_system_generated is True


def test_auditor_role_permissions(session):
    from bayan.api import get_company_legal_archive, create_invoice, InvoiceIn
    from bayan.auth import require, AuthError
    from fastapi import HTTPException
    import pytest
    
    # 1. Verify require function allows auditor to read but not post/approve
    require("auditor", "read")  # should pass without error
    
    with pytest.raises(AuthError):
        require("auditor", "post")
        
    with pytest.raises(AuthError):
        require("auditor", "approve")
        
    # 2. Verify archive export allows admin & auditor but blocks viewer
    c = setup_company(session, "Auditor Co", "retail", is_vat_payer=False)
    
    # Auditor should succeed
    archive = get_company_legal_archive(c.id, ctx={"uid": "user-1", "role": "auditor"}, db=session)
    assert archive["company"]["name"] == "Auditor Co"
    
    # Viewer should fail with HTTP 403
    with pytest.raises(HTTPException) as excinfo:
        get_company_legal_archive(c.id, ctx={"uid": "user-2", "role": "viewer"}, db=session)
    assert excinfo.value.status_code == 403


def test_tax_progressive_pit_and_ndsh_cap(session):
    from bayan.salary import calc_progressive_pit, calc_one, SalaryConfig
    cfg = SalaryConfig(ndsh_ceiling_minor=660000000) # 6,600,000₮ cap
    
    # 1. 5 million ₮ (under 10M limit) -> 10% tax
    # 5,000,000₮ = 500,000,000 minor
    tax_5m = calc_progressive_pit(500000000)
    assert tax_5m == 50000000  # 500,000₮
    
    # 2. 12 million ₮ (between 10M and 15M) -> 1M + 15% on 2M
    # 12,000,000₮ = 1,200,000,000 minor
    tax_12m = calc_progressive_pit(1200000000)
    assert tax_12m == 130000000  # 1,000,000 + 300,000 = 1,300,000₮
    
    # 3. 18 million ₮ (over 15M) -> 1.75M + 20% on 3M
    # 18,000,000₮ = 1,800,000,000 minor
    tax_18m = calc_progressive_pit(1800000000)
    assert tax_18m == 235000000  # 1,750,000 + 600,000 = 2,350,000₮
    
    # 4. NDSH ceiling test: 8,000,000₮ salary (above 6.6M cap)
    c = calc_one(800000000, 22.0, 0, 0, 0, cfg)
    # ndsh_base should be 6.6M, ndsh_emp = 6.6M * 11.5% = 759,000₮
    assert c["ndsh_emp"] == 75900000


def test_bank_dscr_and_indirect_cashflow(session):
    from bayan.models import LoanContract, LoanSchedule
    from bayan.ratios import get_unified_dashboard_summary, get_indirect_cash_flow
    from bayan.ledger import post_entry, LineInput
    
    c = setup_company(session, "Bank Co", "retail", is_vat_payer=False)
    
    # 1. Invest Capital: Cash 2,000,000₮
    post_entry(session, c.id, date(2026, 3, 1), [
        LineInput("1001", debit_minor=200000000),
        LineInput("4101", credit_minor=200000000)
    ])
    # 2. Post Depreciation expense: 20,000₮
    post_entry(session, c.id, date(2026, 3, 5), [
        LineInput("7107", debit_minor=2000000),
        LineInput("2509", credit_minor=2000000)
    ])
    
    # 3. Create Loan and Schedule
    loan = LoanContract(
        company_id=c.id, contract_no="L-99", bank_name="Khan Bank",
        principal_minor=100000000, interest_rate=1.5,
        start_date=date(2026, 3, 1), end_date=date(2026, 12, 31)
    )
    session.add(loan)
    session.flush()
    
    sched = LoanSchedule(
        contract_id=loan.id, due_date=date(2026, 3, 28),
        principal_due_minor=10000000, interest_due_minor=1500000
    )
    session.add(sched)
    session.flush()
    
    # 4. Get ratios and check EBITDA / DSCR / CCC
    dash = get_unified_dashboard_summary(session, c.id, date(2026, 3, 1), date(2026, 3, 31))
    assert "ebitda" in dash["ratios"]
    assert "dscr" in dash["ratios"]
    assert "ccc" in dash["ratios"]
    
    # 5. Verify indirect cash flow
    cf = get_indirect_cash_flow(session, c.id, date(2026, 3, 1), date(2026, 3, 31))
    assert "net_income" in cf
    assert cf["adjustments"]["depreciation"] == 20000.0


def test_salary_overtime_and_vacation(session):
    from bayan.salary import Employee, TimeSheet, run_payroll, PayrollLine
    
    c = setup_company(session, "Salary Co", "retail", is_vat_payer=False)
    e = Employee(company_id=c.id, code="E-OT", last_name="Бат", first_name="Баяр", base_salary_minor=176000000) # 1,760,000₮ (80k daily, 10k hourly)
    session.add(e)
    session.flush()
    
    # 1. 10 hours overtime, 5 hours holiday
    ts = TimeSheet(
        company_id=c.id, employee_id=e.id, year=2026, month=3,
        worked_days=22.0, overtime_hours=10.0, holiday_hours=5.0
    )
    session.add(ts)
    session.flush()
    
    # 2. Run payroll:
    # gross should be: 1.76M (base) + 10 * 10k * 1.5 (150k OT) + 5 * 10k * 2.0 (100k holiday) = 2.01M
    res = run_payroll(session, c.id, 2026, 3)
    assert res["gross"] == 201000000
    
    # 3. Next month (April) vacation pay using 12-month average (March gross is 2.01M)
    ts2 = TimeSheet(
        company_id=c.id, employee_id=e.id, year=2026, month=4,
        worked_days=12.0, vacation_days=10.0  # 10 days vacation
    )
    session.add(ts2)
    session.flush()
    
    res2 = run_payroll(session, c.id, 2026, 4)
    # vacation daily rate should be: 2.01M / 22.0 = 91,363.6₮
    # total gross = 12 * 80k + 10 * 91,363.6₮ = 960k + 913.6k = 1.873M
    assert res2["gross"] >= 187000000


def test_warehouse_reorder_and_transit(session):
    from bayan.inventory import Item, Warehouse, receive, post_stocktake_variance, ship_transfer, receive_transfer, StockMove
    from bayan.ratios import get_unified_dashboard_summary
    
    c = setup_company(session, "Warehouse Co", "retail", is_vat_payer=False)
    wh1 = Warehouse(company_id=c.id, code="WH-1", name="Central")
    wh2 = Warehouse(company_id=c.id, code="WH-2", name="Branch")
    session.add_all([wh1, wh2])
    
    # 1. Create item with safety stock of 10
    item = Item(company_id=c.id, code="I-1", name="Item 1", qty=0, reorder_point=10)
    session.add(item)
    session.flush()
    
    # 2. Check that it triggers reorder alert on dashboard
    dash = get_unified_dashboard_summary(session, c.id, date(2026, 3, 1), date(2026, 3, 31))
    alerts_types = [a["type"] for a in dash["alerts"]]
    assert "INVENTORY_REORDER_ALERT" in alerts_types
    
    # 3. Receive 100 units
    receive(session, c.id, item, date(2026, 3, 1), qty=100, total_cost_minor=1000000, warehouse_id=wh1.id)
    
    # 4. Stocktake variance test: physical count is 98 (deficit of 2)
    post_stocktake_variance(session, c.id, item, 98, date(2026, 3, 5))
    assert item.qty == 98
    
    # 5. Goods in transit test: ship 10 units to WH-2
    ship_move = ship_transfer(session, c.id, item, wh1.id, 10, date(2026, 3, 10))
    # verify it went to transit account 1502
    assert ship_move.target_account == "1502"
    assert "TRANSIT:" in ship_move.ref
    
    # Receive transfer at WH-2
    recv_move = receive_transfer(session, c.id, ship_move.id, wh2.id, date(2026, 3, 12))
    assert recv_move.warehouse_id == wh2.id
    assert "RECEIVED:" in recv_move.ref
















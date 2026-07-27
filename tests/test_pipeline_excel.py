"""End-to-end: синтетик Хаан банкны xlsx → pipeline → gate → журнал.

Бодит хуулга samples/-д ирэхээр эдгээр generator-ийн оронд golden файлууд орно.
"""

from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook

from bayan import ledger
from bayan.coa_seed import add_bank_gl_account
from bayan.models import ClassificationSuggestion, ClassifierRule, BankTxn, Direction
from bayan.pipeline import PipelineError, approve_suggestions, process_file
from bayan.classify import classify_batch


def make_khan_xlsx(path: Path, *, corrupt_balance=False, opening=5_000_000_00):
    """Хаан банкны интернэт банкны хуулгын хэлбэрийг дуурайсан файл."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Дансны хуулга"
    ws["A1"] = "Хаан банк — Дансны хуулга"
    ws["A2"] = "Дансны дугаар:"; ws["B2"] = "5041234567"
    ws["A3"] = "Дансны нэр:";    ws["B3"] = "Тест ХХК"
    ws["A4"] = "Хуулгын хугацаа:"; ws["B4"] = "2026.03.01 - 2026.03.31"
    ws["A5"] = "Эхний үлдэгдэл:"; ws["B5"] = opening / 100
    header = ["Гүйлгээний огноо", "Салбар", "Дебит гүйлгээ", "Кредит гүйлгээ",
              "Үлдэгдэл", "Харьцсан данс", "Гүйлгээний утга"]
    ws.append([]); ws.append(header)

    rows = [
        ("2026.03.02 10:15", "УБ салбар", None, 1_500_000.00, None,
         "5099887766", "Борлуулалтын орлого Т.Бат"),
        ("2026.03.05 14:00", "УБ салбар", 400_000.00, None, None,
         "5011223344", "Түрээсийн төлбөр 3-р сар"),
        ("2026.03.10 09:30", "УБ салбар", 250_000.00, None, None,
         None, "SOCIAL INSURANCE НДШ töлөлт"),
        ("2026.03.15 16:45", "УБ салбар", None, 2_000_000.00, None,
         "5055667788", "Гэрээт ажлын урьдчилгаа"),
        ("2026.03.20 11:00", "УБ салбар", 1_200.00, None, None,
         None, "Гүйлгээний шимтгэл"),
    ]
    balance = opening / 100
    for r in rows:
        balance = balance - (r[2] or 0) + (r[3] or 0)
        ws.append([r[0], r[1], r[2], r[3], balance, r[5], r[6]])

    if corrupt_balance:
        ws.cell(row=10, column=5).value = float(ws.cell(row=10, column=5).value) + 7

    closing = balance if not corrupt_balance else balance  # мета нь зөв хэвээр
    ws.append([])
    ws.append(["Эцсийн үлдэгдэл:", closing])
    ws.append(["Гүйлгээний тоо:", len(rows)])
    wb.save(path)
    return closing


def test_pipeline_end_to_end(session, company, tmp_path):
    f = tmp_path / "khan_statement_2026_03.xlsx"
    make_khan_xlsx(f)

    report = process_file(session, company.id, f)
    assert report.descriptor_id == "khan.internet_bank.xlsx.v1"
    assert report.txn_count == 5
    assert report.gate_ok, report.issues

    txns = session.query(BankTxn).all()
    assert len(txns) == 5
    assert all(t.amount_minor > 0 for t in txns)
    # дүн minor unit-ээр зөв: 1,500,000.00₮ = 150000000 мөнгө
    assert txns[0].amount_minor == 150_000_000
    assert txns[0].direction == Direction.credit
    # утга доторх данс counterparty болж ялгарсан
    assert txns[0].counterparty_account == "5099887766"


def test_pipeline_corrupt_balance_fails_gate_at_row(session, company, tmp_path):
    f = tmp_path / "khan_statement_corrupt.xlsx"
    make_khan_xlsx(f, corrupt_balance=True)

    report = process_file(session, company.id, f)
    assert not report.gate_ok
    v1 = [i for i in report.issues if i["check"] == "V1"]
    assert v1, report.issues
    assert v1[0]["row"] == 10          # яг эвдэрсэн мөрийг заасан
    # gate даваагүй тул гүйлгээ хадгалагдаагүй
    assert session.query(BankTxn).count() == 0


def test_pipeline_duplicate_file_rejected(session, company, tmp_path):
    f = tmp_path / "khan_statement_2026_03.xlsx"
    make_khan_xlsx(f)
    process_file(session, company.id, f)
    with pytest.raises(PipelineError, match="SHA-256"):
        process_file(session, company.id, f)


def test_rules_then_approve_creates_journal(session, company, tmp_path):
    """Дүрмийн ангилалт → батлах → журнал → гүйлгээ баланс банктай таарна."""
    f = tmp_path / "khan_statement_2026_03.xlsx"
    closing = make_khan_xlsx(f)

    bank_gl = add_bank_gl_account(session, company.id, "khan", "5041234567")

    # нябо-гийн дүрмүүд
    session.add_all([
        ClassifierRule(company_id=company.id, keyword="борлуулалтын орлого",
                       account_code="5101", priority=10),
        ClassifierRule(company_id=company.id, keyword="түрээс",
                       account_code="7103", priority=10),
        ClassifierRule(company_id=company.id, keyword="social insurance",
                       account_code="3103", priority=10),
        ClassifierRule(company_id=company.id, keyword="урьдчилгаа",
                       account_code="5105", priority=20),
        ClassifierRule(company_id=company.id, keyword="шимтгэл",
                       account_code="7106", priority=20),
    ])
    session.flush()

    process_file(session, company.id, f)
    txns = session.query(BankTxn).order_by(BankTxn.seq_no).all()
    classify_batch(session, company.id, txns, use_ai=False)

    sugs = session.query(ClassificationSuggestion).all()
    assert len(sugs) == 5
    assert all(s.source == "rule" and s.confidence == 1.0 for s in sugs)

    approve_suggestions(session, company.id, [s.id for s in sugs], bank_gl.code)

    tb = {r["code"]: r for r in ledger.trial_balance(session, company.id)}
    # Банкны GL дансны цэвэр өөрчлөлт = хуулгын цэвэр гүйлгээ (A3 шалгуур)
    net = tb[bank_gl.code]["balance_minor"]
    assert net == int(round((closing - 5_000_000.00) * 100))
    assert tb["5101"]["balance_minor"] == 150_000_000
    assert tb["7103"]["balance_minor"] == 40_000_000

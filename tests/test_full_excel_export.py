"""Бүрэн Excel тайлан — 9 хуудас.

СТ-1..4 нь нэгтгэсэн дүн харуулдаг тул нягтлан бодогч дүн бүрийн ард ямар
данс, харилцагч, бараа байгааг тусад нь татах шаардлагатай болдог. Бүх
дэлгэрэнгүйг нэг файлд гаргаж байгааг энд шалгана.
"""

from datetime import date

import pytest
from openpyxl import load_workbook

from bayan import export_reports, ledger
from bayan.ledger import LineInput


EXPECTED_SHEETS = [
    "СТ-1 Баланс", "СТ-2 Орлого", "СТ-3 Өмч", "СТ-4 Мөнгөн гүйлгээ",
    "5. Гүйлгээ баланс", "6. Авлага-Өглөг", "7. Бараа материал",
    "8. Үндсэн хөрөнгө", "9. НӨАТ (ТТ-03А)",
]


@pytest.fixture
def book(session, company):
    ledger.post_entry(session, company.id, date(2026, 7, 1), [
        LineInput("1101", debit_minor=90_000_00, description="Борлуулалтын орлого"),
        LineInput("5101", credit_minor=90_000_00, description="Борлуулалт"),
    ])
    ledger.post_entry(session, company.id, date(2026, 7, 5), [
        LineInput("7103", debit_minor=20_000_00, description="Түрээс"),
        LineInput("1101", credit_minor=20_000_00, description="Түрээс төлөв"),
    ])
    session.commit()
    path = export_reports.export_excel(session, company.id, 2026, 7)
    wb = load_workbook(path)
    yield wb
    wb.close()
    path.unlink(missing_ok=True)


def test_workbook_has_nine_sheets(book):
    assert book.sheetnames == EXPECTED_SHEETS
    assert len(book.sheetnames) == 9


@pytest.mark.parametrize("title", EXPECTED_SHEETS[4:])
def test_detail_sheets_have_header_and_columns(book, title):
    """Хуудас бүр компанийн нэр, тайлбар, баганын нэртэй байх ёстой."""
    ws = book[title]
    assert ws["A1"].value, f"{title}: компанийн нэр алга"
    assert ws["A2"].value, f"{title}: тайлбар алга"
    assert ws.cell(row=4, column=1).value, f"{title}: баганын нэр алга"
    assert ws.freeze_panes == "A5", f"{title}: толгой мөр царцаагүй"


def test_trial_balance_sheet_lists_used_accounts(book):
    ws = book["5. Гүйлгээ баланс"]
    codes = [ws.cell(row=r, column=1).value for r in range(5, ws.max_row + 1)]
    for code in ("1101", "5101", "7103"):
        assert code in codes, f"{code} данс гүйлгээ балансад алга"


def test_trial_balance_totals_are_equal(book):
    """Гүйлгээний баланс заавал тэнцэнэ."""
    ws = book["5. Гүйлгээ баланс"]
    total = [ws.cell(row=ws.max_row, column=c).value for c in range(1, 6)]
    assert total[1] == "НИЙТ"
    assert total[2] == total[3] == pytest.approx(110_000.00)


def test_vat_sheet_lists_official_form_lines(book):
    ws = book["9. НӨАТ (ТТ-03А)"]
    labels = [ws.cell(row=r, column=1).value for r in range(5, ws.max_row + 1)]
    assert any(l and l.startswith("1.") for l in labels)
    assert any(l and l.startswith("64.") for l in labels)
    assert len(labels) == len(export_reports._TT03A_LABELS)


def test_empty_sections_explain_themselves(book):
    """Өгөгдөлгүй хэсэг хоосон биш, тайлбартай байх ёстой."""
    for title in ("6. Авлага-Өглөг", "7. Бараа материал", "8. Үндсэн хөрөнгө"):
        ws = book[title]
        assert ws.cell(row=5, column=1).value, f"{title}: хоосон тайлбар алга"


def test_export_still_works_with_no_activity(session, company):
    """Гүйлгээ огт байхгүй компанид ч бүрэн файл гарна."""
    path = export_reports.export_excel(session, company.id, 2026, 7)
    wb = load_workbook(path)
    assert wb.sheetnames == EXPECTED_SHEETS
    wb.close()
    path.unlink(missing_ok=True)

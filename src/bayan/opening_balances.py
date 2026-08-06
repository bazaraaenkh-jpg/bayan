"""Эхний үлдэгдэл импортлогч — 1С эсвэл BAAZ системээс экспортолсон Excel үлдэгдлийг уншиж журналын бичилт үүсгэнэ.
"""

from __future__ import annotations

import logging
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any
from sqlalchemy import select
from sqlalchemy.orm import Session

from . import ledger
from .models import Account
from .amounts import parse_amount

# Calamine support
try:
    from python_calamine import CalamineWorkbook
    HAS_CALAMINE = True
except ImportError:
    HAS_CALAMINE = False

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def _clean_code(val: Any) -> str:
    """Дансны кодыг цэвэрлэж цэг болон хоосон зайг арилгана (жишээ нь 1001.01 -> 100101)."""
    if val is None:
        return ""
    text = str(val).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text.replace(".", "").replace(" ", "").strip()


def _parse_detailed_balance_sheet_grid(grid: list[list[Any]], accounts: dict[str, Any]) -> list[ledger.LineInput]:
    import re
    pattern = re.compile(r'^(\d{4,8})\s*[-–—]\s*(.+)$')
    lines: list[ledger.LineInput] = []
    
    is_liability_section = False

    for row in grid:
        if not row:
            continue
        c0 = str(row[0]).strip() if row[0] is not None else ""
        c0_lower = c0.lower()
        
        if "өр төлбөр ба эзэмшигчдийн өмч" in c0_lower or "өр төлбөр болон эздийн өмч" in c0_lower:
            is_liability_section = True
            continue

        if not c0 or c0_lower.startswith("нийт") or c0_lower.startswith("эргэлтийн") or c0_lower.startswith("хөрөнгийн дүн") or c0_lower.startswith("өр төлбөрийн дүн") or c0_lower == "хөрөнгө" or c0_lower == "өр төлбөр":
            continue

        raw_code = None
        desc = ""
        m = pattern.match(c0)
        if m:
            raw_code = m.group(1)
            desc = m.group(2).strip()
        elif "хуримтлагдсан ашиг" in c0_lower:
            raw_code = "4501"
            desc = "Хуримтлагдсан ашиг алдагдал"

        if not raw_code:
            continue

        amt = None
        for cell in reversed(row[1:]):
            if cell is not None and isinstance(cell, (int, float)):
                amt = float(cell)
                break
            elif cell is not None:
                try:
                    txt = str(cell).replace(",", "").replace(" ", "").strip()
                    if txt:
                        amt = float(txt)
                        break
                except ValueError:
                    pass

        if amt is None or amt == 0:
            continue

        target_code = None
        if raw_code in accounts:
            target_code = raw_code
        elif len(raw_code) >= 4 and raw_code[:4] in accounts:
            target_code = raw_code[:4]
        else:
            # Domain fallback mapping
            if raw_code.startswith("14") or raw_code.startswith("20"):
                if raw_code == "2002":
                    target_code = "2509" if "2509" in accounts else "2501" # Хуримтлагдсан элэгдэл
                elif raw_code == "2001":
                    target_code = "2501" if "2501" in accounts else None
                elif raw_code == "1405":
                    target_code = "2101" if "2101" in accounts else "2103"
            elif raw_code.startswith("33"):
                target_code = "3102" if "3102" in accounts else "3101"
            elif raw_code.startswith("34"):
                if raw_code == "3403":
                    target_code = "3106" if "3106" in accounts else "3104"
                else:
                    target_code = "3104" if "3104" in accounts else "3101"
            elif raw_code.startswith("4"):
                if "4501" in accounts:
                    target_code = "4501" if ("ашиг" in desc.lower() or raw_code == "4102") else "4101"
                else:
                    target_code = "4101"
            
            if not target_code:
                prefix2 = raw_code[:2]
                for ac_code in accounts:
                    if ac_code.startswith(prefix2):
                        target_code = ac_code
                        break

        if not target_code:
            continue

        minor_val = parse_amount(abs(amt)) or 0
        if minor_val == 0:
            continue

        if not is_liability_section:
            # Asset section
            if amt > 0:
                lines.append(ledger.LineInput(account_code=target_code, debit_minor=minor_val, description=f"Эхний үлдэгдэл: {desc}"))
            else:
                lines.append(ledger.LineInput(account_code=target_code, credit_minor=minor_val, description=f"Эхний үлдэгдэл: {desc}"))
        else:
            # Liabilities & Equity section
            if amt > 0:
                lines.append(ledger.LineInput(account_code=target_code, credit_minor=minor_val, description=f"Эхний үлдэгдэл: {desc}"))
            else:
                lines.append(ledger.LineInput(account_code=target_code, debit_minor=minor_val, description=f"Эхний үлдэгдэл: {desc}"))

    return lines


def parse_amount_safe(v: Any) -> int:
    if v is None:
        return 0
    try:
        return parse_amount(v) or 0
    except Exception:
        return 0


def _parse_st1_report_grid(grid: list[list[any]], accounts: dict[str, Account]) -> list[ledger.LineInput]:
    ST_MAP = {
        "1.1.1": "1011",
        "1.1.2": "1210",
        "1.1.3": "1310",
        "1.1.4": "1240",
        "1.1.5": "1120",
        "1.1.6": "1510",
        "1.1.7": "1410",
        "1.1.8": "1610",
        "1.2.1": "1810",
        "1.2.2": "1820",
        "1.2.4": "1710",
        "2.1.1": "2110",
        "2.1.1.1": "2110",
        "2.1.1.2": "2130",
        "2.1.1.3": "2120",
        "2.1.1.4": "2120",
        "2.1.1.5": "2150",
        "2.1.1.6": "2140",
        "2.1.2": "2120",
        "2.1.3": "2130",
        "2.3.1": "3010",
        "2.3.2": "3010",
        "2.3.3": "3010",
        "2.3.8": "3110",
        "2.3.9": "3110",
        "2.3.10": "3110",
    }

    def parse_amt_safe(v):
        if v is None: return 0
        try: return parse_amount(v) or 0
        except Exception: return 0

    amt_col = 2
    for r in grid[3:15]:
        c0 = str(r[0]).strip() if len(r) > 0 and r[0] is not None else ""
        if c0 in ST_MAP or (len(c0) >= 3 and c0[0] in "123"):
            for idx, val in enumerate(r[2:], start=2):
                if parse_amt_safe(val) > 0:
                    amt_col = idx
                    break

    current_ac = None
    section_rows = []
    
    for r_idx, row in enumerate(grid):
        if not row:
            continue
        c0 = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        c1 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        amt_raw = row[amt_col] if len(row) > amt_col else None
        amt = parse_amt_safe(amt_raw)

        if "НИЙТ" in c1.upper() or "ДҮН" in c1.upper():
            continue

        if c0 in ST_MAP:
            current_ac = ST_MAP[c0]
            if amt != 0:
                section_rows.append((c0, current_ac, c1, amt, "main"))
        elif current_ac and c1 and not c0:
            if amt != 0:
                section_rows.append(("", current_ac, c1, amt, "sub"))

    final_rows = []
    acs = set(r[1] for r in section_rows)
    for ac in acs:
        sub_items = [r for r in section_rows if r[1] == ac and r[4] == "sub"]
        if sub_items:
            final_rows.extend(sub_items)
        else:
            main_items = [r for r in section_rows if r[1] == ac and r[4] == "main"]
            final_rows.extend(main_items)

    lines = []
    for r in sorted(final_rows, key=lambda x: x[1]):
        code = r[1]
        amt = r[3]
        desc = r[2]
        target_code = code if code in accounts else None
        if not target_code:
            for ac_code in accounts:
                if ac_code.startswith(code[:2]) or ac_code.startswith(code[:1]):
                    target_code = ac_code
                    break
        if not target_code:
            target_code = "1011" if code[0] in ("1", "2") else "2110"

        if code[0] in ("1", "2") and code not in ("2110", "2120", "2130", "2140", "2150"):
            lines.append(ledger.LineInput(account_code=target_code, debit_minor=amt, description=desc or "Эхний үлдэгдэл"))
        else:
            lines.append(ledger.LineInput(account_code=target_code, credit_minor=amt, description=desc or "Эхний үлдэгдэл"))

    return lines


def import_opening_balances(
    session: Session,
    company_id: str,
    file_path: Path,
    opening_date: date,
) -> dict[str, Any]:
    """1С / BAAZ Excel файлаас дансны үлдэгдлийг уншиж эхний үлдэгдлийн журналын бичилт үүсгэнэ."""
    
    # 1. Excel унших (Calamine-ийг илүүд үзэж, openpyxl-ийг fallback болгоно)
    grid: list[list[Any]] = []
    if HAS_CALAMINE:
        try:
            wb = CalamineWorkbook.from_path(str(file_path))
            grid = wb.get_sheet_by_name(wb.sheet_names[0]).to_python()
        except Exception as e:
            logger.warning("Calamine failed to load opening balances file: %s", e)

    if not grid:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            grid = [[cell.value for cell in row] for row in ws.iter_rows()]
        finally:
            wb.close()

    # Одоо байгаа компанийн данснуудын жагсаалт
    accounts = {a.code: a for a in session.scalars(
        select(Account).where(Account.company_id == company_id, Account.is_postable == True)
    ).all()}

    # 2. Толгойн мөрийг олох (Данс/Код, Дебит, Кредит гэсэн үгс орсон)
    header_row_idx = None
    code_col = None
    debit_col = None
    credit_col = None

    code_keywords = ["дансны код", "дансны дугаар", "account code", "код", "code", "дансны №"]
    debit_keywords = ["дебит", "дэбит", "дебет", "дэв", "деб", "дб", "dr", "debit"]
    credit_keywords = ["кредит", "крэдит", "кредет", "кре", "крэ", "кр", "cr", "credit"]

    for r_idx in range(min(50, len(grid))):
        row = grid[r_idx]
        if not row:
            continue
        c0_str = str(row[0]).strip() if row[0] is not None else ""
        # Дижитээр эхэлсэн дата мөрийг толгойн мөр гэж үзэхгүй
        if c0_str and c0_str[0].isdigit():
            continue

        next_row = grid[r_idx + 1] if r_idx + 1 < len(grid) else []

        temp_code = None
        temp_debit = None
        temp_credit = None

        max_cols = max(len(row), len(next_row))
        for c_idx in range(max_cols):
            val1 = str(row[c_idx]).strip().lower() if c_idx < len(row) and row[c_idx] is not None else ""
            val2 = str(next_row[c_idx]).strip().lower() if c_idx < len(next_row) and next_row[c_idx] is not None else ""
            val_str = f"{val1} {val2}".strip()

            if not val_str:
                continue

            is_code_cell = any(k in val_str for k in code_keywords) or ("данс" in val_str and not any(x in val_str for x in ["авлага", "өглөг", "орлого", "зардал"]))

            if temp_code is None and is_code_cell:
                temp_code = c_idx
            if temp_debit is None and any(k in val_str for k in debit_keywords):
                temp_debit = c_idx
            if temp_credit is None and any(k in val_str for k in credit_keywords):
                temp_credit = c_idx

        if temp_code is not None and (temp_debit is not None or temp_credit is not None):
            if temp_code != temp_debit and temp_code != temp_credit:
                header_row_idx = r_idx
                code_col = temp_code
                debit_col = temp_debit
                credit_col = temp_credit
                
                # Position-based fallback if one of debit/credit wasn't explicitly named
                if debit_col is not None and credit_col is None:
                    credit_col = debit_col + 1
                elif credit_col is not None and debit_col is None:
                    debit_col = max(0, credit_col - 1)
                break

    if header_row_idx is None or code_col is None:
        # 3. Ухаалаг авто-дэвсгэрт: дата мөрүүдээс Дансны код болон Дүн бүхий багануудыг олох
        for r_idx in range(min(50, len(grid))):
            row = grid[r_idx]
            if not row:
                continue
            for c_idx, cell in enumerate(row):
                if cell is None:
                    continue
                cell_str = str(cell).strip()
                if not re.match(r'^\d{4,8}(\s*[-–]\s*.*)?$', cell_str):
                    continue
                code_cand = _clean_code(cell)
                if code_cand and code_cand.isdigit() and len(code_cand) in (4, 8) and (code_cand in accounts or code_cand[:4] in accounts):
                    num_cols = []
                    for c2_idx, c2_val in enumerate(row):
                        if c2_idx != c_idx and c2_val is not None:
                            p_amt = parse_amount_safe(c2_val)
                            if p_amt is not None and p_amt != 0:
                                num_cols.append(c2_idx)
                    if num_cols:
                        header_row_idx = max(0, r_idx - 1)
                        code_col = c_idx
                        debit_col = num_cols[0]
                        credit_col = num_cols[1] if len(num_cols) > 1 else num_cols[0]
                        break
            if header_row_idx is not None:
                break

    lines: list[ledger.LineInput] = []
    total_debit = 0
    total_credit = 0

    if header_row_idx is not None and code_col is not None:
        # Standard table parsing
        FALLBACK_MAP = {
            "1001": "1001", "1101": "1011", "1201": "1210", "1204": "1240",
            "1405": "1510", "1501": "1410", "2001": "1810", "2002": "1811",
            "3101": "2110", "3301": "2120", "3401": "2130", "3402": "2130",
            "3403": "2130", "3406": "2140", "4101": "3010", "3110": "3110",
        }

        for r_idx in range(header_row_idx + 1, len(grid)):
            row = grid[r_idx]
            if len(row) <= code_col:
                continue
                
            raw_code = row[code_col]
            code = _clean_code(raw_code)
            if not code or not code.isdigit():
                continue

            target_code = None
            if code in accounts:
                target_code = code
            elif len(code) >= 4 and code[:4] in accounts:
                target_code = code[:4]
            else:
                prefix = code[:4]
                mapped = FALLBACK_MAP.get(prefix) or FALLBACK_MAP.get(code)
                if mapped and mapped in accounts:
                    target_code = mapped
                else:
                    for ac_code in accounts:
                        if ac_code.startswith(code[:2]) or ac_code.startswith(code[:1]):
                            target_code = ac_code
                            break

            if not target_code:
                logger.info("Дансны төлөвлөгөөнд таарахгүй кодыг алгаслаа: %s", code)
                continue

            if debit_col == credit_col:
                raw_amt = row[debit_col] if debit_col < len(row) else None
                amt_val = parse_amount_safe(raw_amt)
                if amt_val != 0:
                    first_c = target_code[0]
                    if first_c in ("1", "2"):
                        if amt_val > 0:
                            lines.append(ledger.LineInput(account_code=target_code, debit_minor=abs(amt_val), description="Эхний үлдэгдэл"))
                            total_debit += abs(amt_val)
                        else:
                            lines.append(ledger.LineInput(account_code=target_code, credit_minor=abs(amt_val), description="Эхний үлдэгдэл"))
                            total_credit += abs(amt_val)
                    else:
                        if amt_val > 0:
                            lines.append(ledger.LineInput(account_code=target_code, credit_minor=abs(amt_val), description="Эхний үлдэгдэл"))
                            total_credit += abs(amt_val)
                        else:
                            lines.append(ledger.LineInput(account_code=target_code, debit_minor=abs(amt_val), description="Эхний үлдэгдэл"))
                            total_debit += abs(amt_val)
            else:
                raw_debit = row[debit_col] if debit_col is not None and debit_col < len(row) else None
                raw_credit = row[credit_col] if credit_col is not None and credit_col < len(row) else None

                debit_val = parse_amount_safe(raw_debit)
                credit_val = parse_amount_safe(raw_credit)

                if debit_val == 0 and credit_val == 0:
                    continue

                if debit_val > 0:
                    lines.append(ledger.LineInput(account_code=target_code, debit_minor=debit_val, description="Эхний үлдэгдэл"))
                    total_debit += debit_val
                if credit_val > 0:
                    lines.append(ledger.LineInput(account_code=target_code, credit_minor=credit_val, description="Эхний үлдэгдэл"))
                    total_credit += credit_val

    if not lines:
        # Fallback 1: Санхүүгийн байдлын тайлан (Маягт СТ-1) форматаар унших
        lines = _parse_st1_report_grid(grid, accounts)
        for line in lines:
            if line.debit_minor:
                total_debit += line.debit_minor
            if line.credit_minor:
                total_credit += line.credit_minor

    if not lines:
        # Fallback 2: Баланс дэлгэрэнгүй (1С / BAAZ / e-Balance) форматаар унших
        lines = _parse_detailed_balance_sheet_grid(grid, accounts)
        for line in lines:
            if line.debit_minor:
                total_debit += line.debit_minor
            if line.credit_minor:
                total_credit += line.credit_minor

    if not lines:
        raise ValueError("Excel файлаас ямар нэгэн дансны үлдэгдэл олдсонгүй.")

    # 4. Дебит болон Кредит тэнцэж буйг шалгана (G1 инвариант)
    if total_debit != total_credit:
        diff = abs(total_debit - total_credit)
        # Хэрэв Excel-ийн бутархайн нарийвчлалын зөрүү 1₮ (100 minor units) хүртэл байвал автоматаар тэнцүүлнэ
        if diff <= 100 and lines:
            if total_debit < total_credit:
                if lines[0].debit_minor is not None and lines[0].debit_minor > 0:
                    lines[0].debit_minor += diff
                else:
                    lines.append(ledger.LineInput(account_code=lines[0].account_code, debit_minor=diff, description="Тооцооны бутархай тэнцүүлэлт"))
                total_debit += diff
            else:
                if lines[0].credit_minor is not None and lines[0].credit_minor > 0:
                    lines[0].credit_minor += diff
                else:
                    lines.append(ledger.LineInput(account_code=lines[0].account_code, credit_minor=diff, description="Тооцооны бутархай тэнцүүлэлт"))
                total_credit += diff

    if total_debit != total_credit:
        diff = abs(total_debit - total_credit)
        raise ValueError(
            f"Баланс зөрүүтэй байна! Нийт Дебит: {total_debit/100:,.2f}₮, "
            f"Нийт Кредит: {total_credit/100:,.2f}₮. Зөрүү: {diff/100:,.2f}₮"
        )

    # 4.5. Хуучин эхний үлдэгдлийн бичилтүүдийг арилгах (Давхардахаас сэргийлнэ)
    from .models import JournalEntry
    old_entries = session.scalars(
        select(JournalEntry).where(
            JournalEntry.company_id == company_id,
            JournalEntry.memo.like("%Эхний үлдэгдэл%")
        )
    ).all()
    for oe in old_entries:
        session.delete(oe)
    session.flush()

    # 5. Гүйлгээг илгээнэ
    entry = ledger.post_entry(
        session, company_id, opening_date, lines,
        source_type=ledger.SourceType.manual,
        memo="Эхний үлдэгдэл импортоор оруулав"
    )

    return {
        "entry_id": entry.id,
        "lines_count": len(lines),
        "total_minor": total_debit,
        "entry_no": entry.entry_no
    }


def import_inventory_from_excel(session: Session, company_id: str, file_path: Path, opening_date: date) -> dict:
    """Excel хуудаснаас бараа болон тэдгээрийн эхний үлдэгдлийг импортлон оруулах."""
    from . import inventory
    
    grid: list[list[Any]] = []
    if HAS_CALAMINE:
        try:
            wb = CalamineWorkbook.from_path(str(file_path))
            grid = wb.get_sheet_by_name(wb.sheet_names[0]).to_python()
        except Exception:
            pass

    if not grid:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            grid = [[cell.value for cell in row] for row in ws.iter_rows()]
        finally:
            wb.close()

    header_row_idx = None
    code_col = None
    name_col = None
    qty_col = None
    cost_col = None
    cur_col = None
    rate_col = None
    duty_col = None
    vat_col = None
    freight_col = None
    date_col = None
    barcode_col = None
    wh_col = None

    for r_idx, row in enumerate(grid[:20]):
        for c_idx, val in enumerate(row):
            if val is None:
                continue
            val_str = str(val).strip().lower()
            if "код" in val_str:
                code_col = c_idx
            elif "нэр" in val_str or "бараа" in val_str:
                name_col = c_idx
            elif "тоо" in val_str or "ширхэг" in val_str or "хэмжээ" in val_str:
                qty_col = c_idx
            elif "өртөг" in val_str or "дүн" in val_str or "үнэ" in val_str:
                cost_col = c_idx
            elif "валют" in val_str:
                cur_col = c_idx
            elif "ханш" in val_str:
                rate_col = c_idx
            elif "гааль" in val_str:
                duty_col = c_idx
            elif "нөат" in val_str or "ноат" in val_str:
                vat_col = c_idx
            elif "тээвэр" in val_str or "нэмэлт" in val_str or "бусад" in val_str:
                freight_col = c_idx
            elif "огноо" in val_str or "date" in val_str:
                date_col = c_idx
            elif "баркод" in val_str or "бар код" in val_str or "barcode" in val_str:
                barcode_col = c_idx
            elif "агуулах" in val_str:
                wh_col = c_idx

        if code_col is not None and name_col is not None:
            header_row_idx = r_idx
            break

    if header_row_idx is None:
        code_col, name_col, qty_col, cost_col = 0, 1, 2, 3
        header_row_idx = 0

    added = 0
    received = 0
    total_vat = 0
    
    default_wh = session.scalar(select(inventory.Warehouse).where(inventory.Warehouse.company_id == company_id))
    if not default_wh:
        default_wh = inventory.Warehouse(company_id=company_id, code="WH01", name="Төв агуулах")
        session.add(default_wh)
        session.flush()

    for row in grid[header_row_idx + 1:]:
        if len(row) <= max(code_col, name_col):
            continue
        code = _clean_code(row[code_col])
        if not code:
            continue
        name = str(row[name_col]).strip() if name_col < len(row) and row[name_col] is not None else f"Бараа {code}"
        
        barcode = str(row[barcode_col]).strip() if barcode_col is not None and barcode_col < len(row) and row[barcode_col] is not None else None
        item = session.scalar(select(inventory.Item).where(inventory.Item.company_id == company_id, inventory.Item.code == code))
        if not item:
            item = inventory.Item(company_id=company_id, code=code, name=name, unit="ш", gl_account="2101", barcode=barcode)
            session.add(item)
            session.flush()
            added += 1
        elif barcode:
            item.barcode = barcode
            
        qty = int(parse_amount(row[qty_col]) / 100) if qty_col is not None and qty_col < len(row) and row[qty_col] is not None else 0
        base_cost = int(parse_amount(row[cost_col])) if cost_col is not None and cost_col < len(row) and row[cost_col] is not None else 0
        
        # Row date
        row_date = opening_date
        if date_col is not None and date_col < len(row) and row[date_col] is not None:
            val = row[date_col]
            if isinstance(val, (date, datetime)):
                row_date = val.date() if hasattr(val, "date") else val
            else:
                try:
                    row_date = date.fromisoformat(str(val).strip()[:10])
                except Exception:
                    pass

        # Currency & exchange rate conversion
        currency = str(row[cur_col]).strip().upper() if cur_col is not None and cur_col < len(row) and row[cur_col] is not None else "MNT"
        rate = float(row[rate_col]) if rate_col is not None and rate_col < len(row) and row[rate_col] is not None else 1.0
        if currency != "MNT" and rate > 0:
            base_cost = int(base_cost * rate)
            
        # Customs duty
        duty = int(parse_amount(row[duty_col])) if duty_col is not None and duty_col < len(row) and row[duty_col] is not None else 0
        
        # Freight / shipping / additional cost
        freight = int(parse_amount(row[freight_col])) if freight_col is not None and freight_col < len(row) and row[freight_col] is not None else 0
        
        # VAT
        vat = int(parse_amount(row[vat_col])) if vat_col is not None and vat_col < len(row) and row[vat_col] is not None else 0
        total_vat += vat
        
        final_cost = base_cost + duty + freight
        
        if qty > 0:
            wh_id = default_wh.id
            if wh_col is not None and wh_col < len(row) and row[wh_col] is not None:
                wh_code = str(row[wh_col]).strip()
                wh = session.scalar(select(inventory.Warehouse).where(inventory.Warehouse.company_id == company_id, inventory.Warehouse.code == wh_code))
                if wh:
                    wh_id = wh.id
                    
            inventory.receive(session, company_id, item, row_date, qty, final_cost, credit_account="3101", ref="Эхний үлдэгдэл импорт (Excel)", warehouse_id=wh_id)
            received += 1

    # Post total import VAT if any
    if total_vat > 0:
        try:
            from . import ledger
            ledger.post_entry(session, company_id, opening_date, [
                ledger.LineInput("1203", debit_minor=total_vat, description="Импортын НӨАТ-ын авлага"),
                ledger.LineInput("3101", credit_minor=total_vat, description="Импортын НӨАТ-ын авлага"),
            ], source_type=ledger.SourceType.manual, memo="Импортын НӨАТ-ын авлага бүртгэв (Excel импорт)")
        except Exception:
            pass

    return {"items_added": added, "balances_imported": received}


def import_assets_from_excel(session: Session, company_id: str, file_path: Path) -> dict:
    """Excel хуудаснаас үндсэн хөрөнгийн жагсаалт ба элэгдлийг бүртгэх."""
    from . import assets
    
    grid: list[list[Any]] = []
    if HAS_CALAMINE:
        try:
            wb = CalamineWorkbook.from_path(str(file_path))
            grid = wb.get_sheet_by_name(wb.sheet_names[0]).to_python()
        except Exception:
            pass

    if not grid:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            grid = [[cell.value for cell in row] for row in ws.iter_rows()]
        finally:
            wb.close()

    header_row_idx = None
    code_col = None
    name_col = None
    cost_col = None
    life_col = None
    date_col = None

    for r_idx, row in enumerate(grid[:20]):
        for c_idx, val in enumerate(row):
            if val is None:
                continue
            val_str = str(val).strip().lower()
            if "код" in val_str:
                code_col = c_idx
            elif "нэр" in val_str or "хөрөнгө" in val_str:
                name_col = c_idx
            elif "өртөг" in val_str or "дүн" in val_str or "үнэ" in val_str:
                cost_col = c_idx
            elif "хугацаа" in val_str or "сар" in val_str:
                life_col = c_idx
            elif "огноо" in val_str or "ашиглалтад" in val_str:
                date_col = c_idx

        if code_col is not None and name_col is not None:
            header_row_idx = r_idx
            break

    if header_row_idx is None:
        code_col, name_col, cost_col, life_col = 0, 1, 2, 3
        header_row_idx = 0

    added = 0
    for row in grid[header_row_idx + 1:]:
        if len(row) <= max(code_col, name_col):
            continue
        code = _clean_code(row[code_col])
        if not code:
            continue
        name = str(row[name_col]).strip() if name_col < len(row) and row[name_col] is not None else f"Хөрөнгө {code}"
        
        from .assets import FixedAsset
        asset = session.scalar(select(FixedAsset).where(FixedAsset.company_id == company_id, FixedAsset.code == code))
        if not asset:
            cost = int(parse_amount(row[cost_col])) if cost_col is not None and cost_col < len(row) and row[cost_col] is not None else 0
            life = int(parse_amount(row[life_col]) / 100) if life_col is not None and life_col < len(row) and row[life_col] is not None else 36
            if life <= 0:
                life = 36
            
            in_service_date = date.today()
            if date_col is not None and date_col < len(row) and row[date_col] is not None:
                try:
                    in_service_date = date.fromisoformat(str(row[date_col]).strip().split(" ")[0])
                except Exception:
                    pass
                    
            assets.register_asset(session, company_id, code, name, cost, life, in_service_date)
            added += 1

    return {"assets_added": added}


def import_inventory_issue_from_excel(session, company_id: str, file_path: Path) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    grid = []
    for r in ws.iter_rows(values_only=True):
        grid.append(list(r))
    wb.close()
    
    header_row_idx = None
    date_col = None
    code_col = None
    qty_col = None
    target_col = None
    wh_col = None
    
    for r_idx, row in enumerate(grid[:20]):
        for c_idx, val in enumerate(row):
            if val is None:
                continue
            val_str = str(val).strip().lower()
            if "огноо" in val_str or "date" in val_str:
                date_col = c_idx
            elif "код" in val_str:
                code_col = c_idx
            elif "тоо" in val_str or "ширхэг" in val_str or "хэмжээ" in val_str:
                qty_col = c_idx
            elif "данс" in val_str or "харьцах" in val_str:
                target_col = c_idx
            elif "агуулах" in val_str:
                wh_col = c_idx
                
        if code_col is not None and qty_col is not None:
            header_row_idx = r_idx
            break
            
    if header_row_idx is None:
        date_col, code_col, qty_col, target_col = 0, 1, 2, 3
        header_row_idx = 0
        
    issued_count = 0
    total_cost_minor = 0
    
    from . import inventory
    
    for row in grid[header_row_idx + 1:]:
        if len(row) <= max(code_col, qty_col):
            continue
            
        code = _clean_code(row[code_col])
        if not code:
            continue
            
        item = session.scalar(select(inventory.Item).where(
            inventory.Item.company_id == company_id,
            inventory.Item.code == code
        ))
        if not item:
            continue
            
        qty = int(parse_amount(row[qty_col]) / 100) if qty_col is not None and qty_col < len(row) and row[qty_col] is not None else 0
        if qty <= 0:
            continue
            
        target_account = "6101"
        if target_col is not None and target_col < len(row) and row[target_col] is not None:
            target_account = str(row[target_col]).strip().split(" ")[0]
            
        # Parse date
        move_date = date.today()
        if date_col is not None and date_col < len(row) and row[date_col] is not None:
            val = row[date_col]
            if isinstance(val, (date, datetime)):
                move_date = val.date() if hasattr(val, "date") else val
            else:
                try:
                    move_date = date.fromisoformat(str(val).strip().split(" ")[0])
                except Exception:
                    pass
                    
        warehouse_code = None
        if wh_col is not None and wh_col < len(row) and row[wh_col] is not None:
            warehouse_code = str(row[wh_col]).strip()
            
        warehouse_id = None
        if warehouse_code:
            wh = session.scalar(select(inventory.Warehouse).where(
                inventory.Warehouse.company_id == company_id,
                inventory.Warehouse.code == warehouse_code
            ))
            if wh:
                warehouse_id = wh.id
                
        try:
            mv = inventory.issue(session, company_id, item, move_date, qty, target_account, warehouse_id=warehouse_id)
            issued_count += 1
            total_cost_minor += mv.cost_minor
        except Exception:
            pass
            
    return {"issued_count": issued_count, "total_cost_minor": total_cost_minor}



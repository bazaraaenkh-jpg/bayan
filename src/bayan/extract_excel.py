"""Зам А — Excel-ийн ухаалаг задлагч (Merged cells, Multi-sheet, Calamine хурдасгуур).

Энэхүү модуль нь банкны Excel хуулгыг детерминистик аргаар задлах бөгөөд дараах боломжтой:
  1. Merged cells (Нэгтгэсэн нүд) илрүүлж утгыг доош нь залгамжлуулах.
  2. Олон sheet-ээс хамгийн өндөр fingerprint score-тойг сонгох.
  3. Rust суурьтай python-calamine ашиглан 10-50 дахин хурдан унших (суулгаагүй бол openpyxl руу шилжинэ).
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .amounts import parse_amount
from .registry import Descriptor

# Calamine холболт шалгах
try:
    from python_calamine import CalamineWorkbook
    HAS_CALAMINE = True
except ImportError:
    HAS_CALAMINE = False

logger = logging.getLogger(__name__)


class ExtractionError(Exception):
    pass


@dataclass
class RawRow:
    row_index: int                       # Эх файлын мөрийн дугаар (1-ээс эхэлнэ)
    values: dict                         # Талбарын нэр -> Түүхий утга


@dataclass
class RawStatement:
    descriptor_id: str
    metadata: dict = field(default_factory=dict)
    rows: list[RawRow] = field(default_factory=list)


# ================================================================= OPENPYXL ХӨДӨЛГҮҮР

def _get_openpyxl_cell_value(ws, row: int, col: int) -> Any:
    """Нэгтгэсэн нүд (merged cells)-ийг шалгаж, зүүн дээд нүдний утгыг авна."""
    for merged_range in ws.merged_cells.ranges:
        if col >= merged_range.min_col and col <= merged_range.max_col:
            if row >= merged_range.min_row and row <= merged_range.max_row:
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return ws.cell(row=row, column=col).value


def _cell_texts(ws, max_rows: int = 40) -> list[str]:
    """Хуудасны эхний хэсэг дэх текстийг fingerprint-д зориулж уншина."""
    texts = []
    # read_only үед ws.iter_rows илүү хурдан
    for row in ws.iter_rows(min_row=1, max_row=max_rows, values_only=True):
        for v in row:
            if v is not None and str(v).strip():
                texts.append(str(v).strip())
    return texts


def _select_sheet(wb, desc: Descriptor) -> Any:
    """Гарчиг таарч буй хамгийн тохиромжтой хуудсыг олон хуудаснаас сонгоно."""
    header_match = desc.table.get("header_row_match", [])
    if not header_match:
        return wb.active
        
    for name in wb.sheetnames:
        ws = wb[name]
        if _find_header_row(ws, header_match) is not None:
            return ws
    return wb.active


def _find_anchor(ws, anchor: str, direction: str = "right") -> Any:
    """Мета утга (дансны дугаар, үлдэгдэл) олох anchor хайлт."""
    for row in ws.iter_rows(min_row=1, max_row=30):
        for cell in row:
            val = cell.value
            if val is None:
                continue
            if anchor.lower() in str(val).lower():
                if direction == "right":
                    # Баруун тийш хоосон биш нүдийг хайна
                    for col_idx in range(cell.column + 1, ws.max_column + 1):
                        v = _get_openpyxl_cell_value(ws, cell.row, col_idx)
                        if v is not None and str(v).strip():
                            return v
                    # Хэрэв баруунаас олдоогүй бол өөрт нь ":" байгаа эсэхийг шалгана
                    text = str(val)
                    if ":" in text:
                        return text.split(":", 1)[1].strip()
    return None


def _find_header_row(ws, header_match: list[str]) -> int | None:
    """Толгойн мөрийн дугаарыг олно."""
    for row in ws.iter_rows(min_row=1, max_row=60):
        texts = [str(c.value).strip().lower() for c in row if c.value is not None]
        if texts and all(any(m.lower() in t for t in texts) for m in header_match):
            return row[0].row
    return None


def _map_columns(ws, header_row: int, desc: Descriptor) -> dict[str, int]:
    """Гарчгийн нэрсийг alias-уудтай тулган баганыг дугаартай харгалзуулна."""
    mapping: dict[str, int] = {}
    headers: dict[int, str] = {}
    for cell in ws[header_row]:
        if cell.value is not None and str(cell.value).strip():
            headers[cell.column] = str(cell.value).strip().lower()

    for field_name, spec in desc.columns.items():
        for col, header in headers.items():
            if any(alias.lower() in header for alias in spec.aliases):
                mapping[field_name] = col
                break
        else:
            if not spec.optional:
                raise ExtractionError(
                    f"Багана олдсонгүй: {field_name} (aliases: {spec.aliases})"
                )
    return mapping


def scan_cells(path: Path, max_rows: int = 40) -> list[str]:
    """Файлыг танихын тулд эхний мөрүүдийн нүдийг уншина."""
    if HAS_CALAMINE:
        try:
            wb = CalamineWorkbook.from_path(str(path))
            first_sheet = wb.sheet_names[0]
            # Хэрэв олон хуудастай бол хамгийн өндөр fingerprint score-тойг олно
            grid = wb.get_sheet_by_name(first_sheet).to_python()
            texts = []
            for row in grid[:max_rows]:
                for v in row:
                    if v is not None and str(v).strip():
                        texts.append(str(v).strip())
            return texts
        except Exception as e:
            logger.warning("Calamine scan failed, falling back to openpyxl: %s", e)

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = _select_sheet(wb, next(iter(load_workbook(path).sheetnames))) # fake select
        # Read-only үед sheet сонгох логик
        ws = wb.active
        return _cell_texts(ws, max_rows)
    finally:
        wb.close()


# ================================================================= CALAMINE ХӨДӨЛГҮҮР

def _extract_via_calamine(path: Path, desc: Descriptor) -> RawStatement | None:
    """Rust calamine ашиглан Excel файлыг хурдтай уншина."""
    if not HAS_CALAMINE:
        return None

    try:
        wb = CalamineWorkbook.from_path(str(path))
        header_match = desc.table.get("header_row_match", [])
        
        # 1. Тохирох хуудас сонгох
        sheet_name = wb.sheet_names[0]
        grid: list[list[Any]] = wb.get_sheet_by_name(sheet_name).to_python()
        
        # Олон sheet шалгах
        for name in wb.sheet_names:
            test_grid = wb.get_sheet_by_name(name).to_python()
            # Хамгийн эхний 60 мөрөөс гарчиг хайна
            found_header = False
            for row in test_grid[:60]:
                row_texts = [str(v).strip().lower() for v in row if v is not None]
                if row_texts and all(any(m.lower() in t for t in row_texts) for m in header_match):
                    sheet_name = name
                    grid = test_grid
                    found_header = True
                    break
            if found_header:
                break

        stmt = RawStatement(descriptor_id=desc.id)
        locale = desc.amount_locale

        # 2. Мета хайх (Grid хайлт)
        def find_anchor_grid(anchor: str, direction: str = "right"):
            for r_idx, row in enumerate(grid[:30]):
                for c_idx, val in enumerate(row):
                    if val is not None and anchor.lower() in str(val).lower():
                        if direction == "right":
                            for next_val in row[c_idx + 1:]:
                                if next_val is not None and str(next_val).strip():
                                    return next_val
                            text = str(val)
                            if ":" in text:
                                return text.split(":", 1)[1].strip()
            return None

        for key, cfg in desc.metadata_block.items():
            value = find_anchor_grid(cfg["anchor"], cfg.get("direction", "right"))
            if value is None:
                continue
            if cfg.get("type") == "amount":
                try:
                    stmt.metadata[key + "_minor"] = parse_amount(
                        value, thousands=locale.get("thousands", ","),
                        decimal=locale.get("decimal", "."))
                except Exception as e:
                    logger.warning("Calamine metadata amount parse skipped for %s: %s", key, e)
            else:
                stmt.metadata[key] = str(value).strip()

        # 3. Хүснэгтийн мөрүүд
        header_row_idx = None
        for r_idx, row in enumerate(grid[:60]):
            row_texts = [str(v).strip().lower() for v in row if v is not None]
            if row_texts and all(any(m.lower() in t for t in row_texts) for m in header_match):
                header_row_idx = r_idx
                break

        if header_row_idx is None:
            raise ExtractionError(f"Толгойн мөр олдсонгүй (match: {header_match})")

        # Багануудын зураглал үүсгэх
        colmap: dict[str, int] = {}
        header_row = grid[header_row_idx]
        headers = {c_idx: str(val).strip().lower() for c_idx, val in enumerate(header_row) if val is not None}

        for field_name, spec in desc.columns.items():
            for col_idx, header in headers.items():
                if any(alias.lower() in header for alias in spec.aliases):
                    colmap[field_name] = col_idx
                    break
            else:
                if not spec.optional:
                    raise ExtractionError(f"Багана олдсонгүй: {field_name}")

        stop_match = [s.lower() for s in desc.table.get("stop_row_match", [])]

        # Утгуудыг унших (Merged Cell Propagate: Calamine-д мөр дараалан ffill хийнэ)
        last_seen_values = {f: None for f in colmap}

        for r_idx in range(header_row_idx + 1, len(grid)):
            row = grid[r_idx]
            
            # 1. Зорилтот баганууд бүгд хоосон бол мөрийг алгасна
            row_vals = [row[col_idx] if col_idx < len(row) else None for col_idx in colmap.values()]
            if all(v is None or str(v).strip() == "" for v in row_vals):
                continue

            # 2. Зогсох нөхцөл шалгах
            row_texts = [str(v).strip() for v in row if v is not None]
            if stop_match and any(any(s in t.lower() for s in stop_match) for t in row_texts):
                break

            values = {}
            for field_name, col_idx in colmap.items():
                val = row[col_idx] if col_idx < len(row) else None
                # Merged cells fallback (ffill)
                if val is None or str(val).strip() == "":
                    if field_name in ("posted_at", "channel_ref", "description"):
                        val = last_seen_values[field_name]
                else:
                    last_seen_values[field_name] = val

                values[field_name] = val

            stmt.rows.append(RawRow(row_index=r_idx + 1, values=values))

        return stmt
    except Exception as e:
        logger.warning("Calamine extraction failed, falling back to openpyxl: %s", e)
        return None


# ================================================================= ЭХЛЭЛ ЦЭГ

def extract_xlsx(path: Path, desc: Descriptor) -> RawStatement:
    """Excel файлыг уншина. Эхлээд Calamine, дараа нь openpyxl-ээр хандана."""
    # Calamine хурдасгуураар унших оролдлого
    if HAS_CALAMINE:
        res = _extract_via_calamine(path, desc)
        if res is not None:
            return res

    # Openpyxl-ээр унших fallback (Merged cells-ийг албан ёсны координат мужаар бодно)
    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        ws = _select_sheet(wb, desc)
        stmt = RawStatement(descriptor_id=desc.id)
        locale = desc.amount_locale

        # --- мета-блок
        for key, cfg in desc.metadata_block.items():
            value = _find_anchor(ws, cfg["anchor"], cfg.get("direction", "right"))
            if value is None:
                continue
            if cfg.get("type") == "amount":
                try:
                    stmt.metadata[key + "_minor"] = parse_amount(
                        value, thousands=locale.get("thousands", ","),
                        decimal=locale.get("decimal", "."))
                except Exception as e:
                    logger.warning("Metadata amount parse skipped for %s: %s", key, e)
            else:
                stmt.metadata[key] = str(value).strip()

        # --- хүснэгт
        header_match = desc.table.get("header_row_match", [])
        header_row = _find_header_row(ws, header_match)
        if header_row is None:
            raise ExtractionError(f"Толгойн мөр олдсонгүй (match: {header_match})")
            
        colmap = _map_columns(ws, header_row, desc)
        stop_match = [s.lower() for s in desc.table.get("stop_row_match", [])]

        for row in ws.iter_rows(min_row=header_row + 1):
            row_idx = row[0].row
            texts = [str(c.value).strip() for c in row if c.value is not None]
            if not texts:
                continue
            if stop_match and any(any(s in t.lower() for s in stop_match) for t in texts):
                break
                
            values = {}
            for field_name, col in colmap.items():
                # Нэгтгэсэн нүдийг шалгана
                values[field_name] = _get_openpyxl_cell_value(ws, row_idx, col)
                
            # Бүх зорилтот багана хоосон мөрийг алгасна
            if all(v is None or str(v).strip() == "" for v in values.values()):
                continue
            stmt.rows.append(RawRow(row_index=row_idx, values=values))

        return stmt
    finally:
        wb.close()

"""Ebarimt / НӨАТУС холболтын албан ёсны клиент.

  * Борлуулалтын баримт үүсгэх (хувь хүн болон байгууллагад)
  * Баримт буцаах (хүчингүй болгох)
  * Худалдан авалтын падаануудыг ТЕГ-ын системээс автоматаар татах
  * Тест болон локал горимд Mock хариу буцаах дэмжлэгтэй.
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field
from typing import Any
import logging

logger = logging.getLogger(__name__)


@dataclass
class EbarimtConfig:
    base_url: str = os.environ.get("EBARIMT_URL", "https://api.ebarimt.mn")
    tin: str = os.environ.get("EBARIMT_TIN", "")
    token: str = os.environ.get("EBARIMT_TOKEN", "")


class EbarimtClient:
    def __init__(self, cfg: EbarimtConfig | None = None):
        self.cfg = cfg or EbarimtConfig()
        # Хэрэв токен нь "mock" эсвэл хоосон бол туршилтын горимд ажиллана
        self.is_mock = self.cfg.token == "mock" or not (self.cfg.tin and self.cfg.token)

    def fetch_receipts(self, year: int, month: int) -> list[dict]:
        """Тухайн сарын байгууллагын борлуулалтын e-barimts татах."""
        if self.is_mock:
            # Mock дата буцаана
            return [
                {"date": f"{year}-{month:02d}-05", "total_minor": 220000000, "receipt_id": "MOCK-REC-01"},
                {"date": f"{year}-{month:02d}-10", "total_minor": 5500000, "receipt_id": "MOCK-REC-02"}
            ]

        import httpx
        r = httpx.get(
            f"{self.cfg.base_url}/receipts",
            params={"tin": self.cfg.tin, "year": year, "month": month},
            headers={"Authorization": f"Bearer {self.cfg.token}"},
            timeout=60
        )
        r.raise_for_status()
        return [
            {"date": x["date"][:10],
             "total_minor": int(round(float(x["totalAmount"]) * 100)),
             "receipt_id": x["id"]}
            for x in r.json().get("receipts", [])
        ]

    def create_receipt(self, amount_minor: int, vat_minor: int,
                       customer_tin: str | None = None,
                       items: list[dict] | None = None) -> dict:
        """Борлуулалтын баримт үүсгэнэ (НӨАТУС API).

        amount_minor: нийт дүн (мөнгөөр)
        vat_minor: НӨАТ-ын дүн (мөнгөөр)
        customer_tin: Байгууллагын РД (байгууллагын баримт бол)
        items: барааны дэлгэрэнгүй жагсаалт: [{"code": "...", "name": "...", "qty": 1, "price_minor": 100}]

        Буцаах утга: {"receipt_id": str, "qr_data": str, "lottery": str, "success": bool}
        """
        amount_tug = float(amount_minor) / 100.0
        vat_tug = float(vat_minor) / 100.0

        if self.is_mock:
            # Mock баримт үүсгэх
            import uuid
            import random
            receipt_id = f"MOCK-REC-{uuid.uuid4().hex[:8].upper()}"
            lottery_code = f"LOT-{random.randint(10000000, 99999999)}"
            return {
                "receipt_id": receipt_id,
                "qr_data": f"https://ebarimt.mn/qr/{receipt_id}",
                "lottery": lottery_code,
                "success": True
            }

        import httpx
        payload = {
            "amount": amount_tug,
            "vat": vat_tug,
            "cashAmount": amount_tug,
            "nonCashAmount": 0.0,
            "billType": "3" if customer_tin else "1",
            "customerTin": customer_tin or "",
            "stocks": []
        }

        if items:
            for it in items:
                price_tug = float(it["price_minor"]) / 100.0
                qty = float(it["qty"])
                payload["stocks"].append({
                    "code": it["code"],
                    "name": it["name"],
                    "qty": qty,
                    "price": price_tug,
                    "vat": price_tug * qty * 0.1, # 10% VAT
                    "totalAmount": price_tug * qty
                })

        r = httpx.post(
            f"{self.cfg.base_url}/receipt/create",
            json=payload,
            headers={"Authorization": f"Bearer {self.cfg.token}"},
            timeout=60
        )
        r.raise_for_status()
        res = r.json()
        
        return {
            "receipt_id": res.get("billId"),
            "qr_data": res.get("qrData"),
            "lottery": res.get("lottery"),
            "success": res.get("success", False)
        }

    def void_receipt(self, receipt_id: str) -> dict:
        """Баримтыг буцааж хүчингүй болгоно."""
        if self.is_mock:
            return {"receipt_id": receipt_id, "success": True}

        import httpx
        r = httpx.post(
            f"{self.cfg.base_url}/receipt/void",
            json={"billId": receipt_id},
            headers={"Authorization": f"Bearer {self.cfg.token}"},
            timeout=60
        )
        r.raise_for_status()
        res = r.json()
        return {
            "receipt_id": receipt_id,
            "success": res.get("success", False)
        }

    def fetch_purchase_invoices(self, year: int, month: int) -> list[dict]:
        """Байгууллагын регистрт ТЕГ-ын системээс ирүүлсэн худалдан авалтын падаануудыг автоматаар татна.

        Буцаах бүтэц: [{"date": "YYYY-MM-DD", "total_minor": int, "vat_minor": int,
                        "supplier_tin": str, "supplier_name": str, "invoice_id": str}]
        """
        if self.is_mock:
            # Mock худалдан авалтын падаанууд
            return [
                {
                    "date": f"{year}-{month:02d}-08",
                    "total_minor": 132000000, # 1,320,000₮ (120,000₮ НӨАТ орсон)
                    "vat_minor": 12000000,
                    "supplier_tin": "5011223344",
                    "supplier_name": "Хангамж Групп ХХК",
                    "invoice_id": "INV-MOCK-P1"
                }
            ]

        import httpx
        r = httpx.get(
            f"{self.cfg.base_url}/purchase-invoices",
            params={"tin": self.cfg.tin, "year": year, "month": month},
            headers={"Authorization": f"Bearer {self.cfg.token}"},
            timeout=60
        )
        r.raise_for_status()
        return [
            {
                "date": x["date"][:10],
                "total_minor": int(round(float(x["totalAmount"]) * 100)),
                "vat_minor": int(round(float(x["vatAmount"]) * 100)),
                "supplier_tin": x["supplierTin"],
                "supplier_name": x["supplierName"],
                "invoice_id": x["invoiceId"]
            }
            for x in r.json().get("invoices", [])
        ]

# =====================================================================
#  eBarimt экспорт файлын задлалт (4 төрлийн тайлан)
# =====================================================================
#
# НӨАТУС-ын хувийн кабинетаас сар бүр 4 тайлан татагдана:
#
#   1. Байгууллагын орлого  — ААН-д борлуулсан   → банкинд мөнгө ОРНО
#   2. Байгууллагын зарлага — ААН-аас худалдсан  → банкнаас мөнгө ГАРНА
#   3. Иргэний орлого       — иргэнд борлуулсан  → банкинд мөнгө ОРНО
#   4. Иргэний зарлага      — иргэний нэрээр худалдсан → банкнаас ГАРНА
#
# Эдгээрийг банкны хуулгатай тулгахад чиглэл (орлого/зарлага) заавал таарах
# ёстой — эс тэгвэл зарлагын падаан орлогын гүйлгээтэй санамсаргүй
# «тулгагдана».

import io
import re
from datetime import date as _date, datetime as _datetime, timedelta as _timedelta
from decimal import Decimal, InvalidOperation


class EbarimtParseError(ValueError):
    """Файлын бүтэц танигдаагүй — хэрэглэгчид юу дутуу байгааг хэлнэ."""


#: dataset түлхүүр → (харагдах нэр, чиглэл, харилцагчийн төрөл)
DATASET_META: dict[str, dict] = {
    "org_income":      {"label": "Байгууллагын орлого",  "direction": "in",  "party_kind": "org"},
    "org_expense":     {"label": "Байгууллагын зарлага", "direction": "out", "party_kind": "org"},
    "citizen_income":  {"label": "Иргэний орлого",       "direction": "in",  "party_kind": "citizen"},
    "citizen_expense": {"label": "Иргэний зарлага",      "direction": "out", "party_kind": "citizen"},
}

_KIND_ORG = ("байгууллаг", "аан", "b2b", "org", "company", "хуулийн этгээд")
_KIND_CITIZEN = ("иргэн", "хувь хүн", "b2c", "citizen", "individual")
_FLOW_IN = ("орлого", "борлуулалт", "борлуулсан", "sale", "income", "revenue")
_FLOW_OUT = ("зарлага", "худалдан авалт", "худалдан", "авалт", "purchase", "expense")


def detect_dataset(*hints: str | None) -> str | None:
    """Файлын нэр / толгойн текстээс 4 тайлангийн алийг нь таних.

    Хоёр тэнхлэг тус тусдаа тодорхойлогдоно: харилцагчийн төрөл (ААН эсвэл
    иргэн) ба мөнгөний чиглэл (орлого эсвэл зарлага). Хоёулаа тодорхой
    болсон үед л dataset буцаана — таамаглахгүй.
    """
    kind = flow = None
    # Дараалал нь чухал: файлын нэр толгойн мөрнөөс давамгайлна. Иргэний
    # тайлангийн толгойд ч «Байгууллагын нэр» гэсэн багана байдаг тул
    # бүх текстийг нэг дор шалгавал буруу таних эрсдэлтэй.
    for hint in hints:
        if not hint:
            continue
        text = hint.lower().replace("_", " ").replace("-", " ")
        if kind is None:
            if any(k in text for k in _KIND_ORG):
                kind = "org"
            elif any(k in text for k in _KIND_CITIZEN):
                kind = "citizen"
        if flow is None:
            if any(k in text for k in _FLOW_OUT):
                flow = "out"
            elif any(k in text for k in _FLOW_IN):
                flow = "in"
        if kind and flow:
            break
    if kind is None or flow is None:
        return None
    for key, meta in DATASET_META.items():
        if meta["party_kind"] == kind and meta["direction"] == flow:
            return key
    return None


# ------------------------------------------------------- мөр унших (3 формат)

def _read_rows(data: bytes, filename: str | None = None) -> list[list]:
    """Файлын төрлийг агуулгаар нь таньж, нүднүүдийг ТҮҮХИЙ утгаар буцаана.

    Огноо, тоог мөр болгож хувиргахгүй — Excel-ийн datetime/тооны нүдийг
    текст болгосноор задлалтын алдаа үүсдэг тул хөрвүүлэлтийг доор нь
    нэг дор хийнэ.
    """
    if not data:
        return []

    # xlsx / xlsm — ZIP (PK)
    if data[:2] == b"PK":
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        best: list[list] = []
        for ws in wb.worksheets:
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            if len(rows) > len(best):
                best = rows
        return best

    # хуучин .xls — OLE2 (BIFF). openpyxl эдгээрийг уншиж чадахгүй тул
    # өмнө нь CSV гэж үзээд хог өгөгдөл үүсгэдэг байв.
    if data[:4] == b"\xd0\xcf\x11\xe0":
        import xlrd
        from xlrd.xldate import xldate_as_datetime
        wb = xlrd.open_workbook(file_contents=data)
        best = []
        for ws in wb.sheets():
            rows = []
            for r in range(ws.nrows):
                row = []
                for c in range(ws.ncols):
                    cell = ws.cell(r, c)
                    if cell.ctype == 3:            # XL_CELL_DATE
                        try:
                            row.append(xldate_as_datetime(cell.value, wb.datemode))
                            continue
                        except Exception:
                            pass
                    row.append(cell.value)
                rows.append(row)
            if len(rows) > len(best):
                best = rows
        return best

    # CSV / TSV
    import csv
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1251", "cp1252"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = data.decode("utf-8", errors="replace")
    sample = text[:4096]
    delim = ","
    try:
        delim = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except Exception:
        for cand in (";", "\t", "|"):
            if sample.count(cand) > sample.count(","):
                delim = cand
    return [list(r) for r in csv.reader(io.StringIO(text), delimiter=delim)]


# ------------------------------------------------------- баганын таних логик

#: талбар → ([(түлхүүр үг, оноо)], [хориглох үг])
_FIELD_RULES: dict[str, tuple[list[tuple[str, int]], list[str]]] = {
    "date": ([("баримтын огноо", 12), ("гүйлгээний огноо", 12), ("огноо", 10),
              ("date", 8), ("он сар", 6), ("өдөр", 4)],
             ["хэвлэ", "илгээ", "тайлант"]),
    "total": ([("нийт дүн", 14), ("нийт үнэ", 14), ("төлбөрийн дүн", 13),
               ("нийт төлбөр", 13), ("нийт", 8), ("total", 8),
               ("дүн", 6), ("amount", 6), ("үнэ", 4)],
              ["нөат", "нхат", "vat", "хот", "city", "хөнгөл", "хямдр",
               "бонус", "татвар", "үлдэгдэл", "тоо ширхэг"]),
    "vat": ([("нөат", 12), ("vat", 10)], ["хувь", "хасах"]),
    "city_tax": ([("нхат", 12), ("citytax", 10), ("хот", 6)], []),
    "ddtd": ([("ддтд", 14), ("ddtd", 14), ("баримтын дугаар", 12),
              ("билл", 8), ("bill", 8), ("баримт", 7), ("дугаар", 4), ("id", 3)],
             ["татвар төлөгч", "регистр", "тин", "tin", "дансны", "утасны"]),
    "party_tin": ([("татвар төлөгчийн дугаар", 14), ("регистр", 11),
                   ("ттд", 9), ("tin", 9), ("рд", 4)], []),
    "party": ([("байгууллагын нэр", 14), ("нийлүүлэгчийн нэр", 14),
               ("харилцагч", 12), ("нийлүүлэгч", 12), ("худалдагч", 12),
               ("худалдан авагч", 12), ("хэрэглэгч", 10), ("merchant", 9),
               ("нэр", 6)],
              ["бараа", "үйлчилгээ", "хэмжих", "нэгж", "файл", "салбар"]),
    "status": ([("төлөв", 10), ("status", 8)], []),
}

_VOID_WORDS = ("хүчингүй", "цуцл", "буцаа", "void", "cancel")


def _cell_text(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (_datetime, _date)):
        return v.isoformat()
    return str(v).strip()


def _score_header_cell(text: str) -> dict[str, int]:
    low = " ".join(text.lower().split())
    if not low:
        return {}
    out: dict[str, int] = {}
    for field, (kws, negs) in _FIELD_RULES.items():
        if any(n in low for n in negs):
            continue
        best = 0
        for kw, w in kws:
            if kw in low:
                best = max(best, w)
        if best:
            out[field] = best
    return out


def _map_columns(header: list) -> dict[str, int]:
    """Багана ↔ талбарын хамгийн өндөр оноотой хослолыг шуналтайгаар сонгоно.

    Өмнөх хувилбар нь СҮҮЛД таарсан баганаар дардаг байсан тул «Нийт дүн»-г
    «НӨАТ-ын дүн» дарж, НӨАТ-ын дүнг нийт дүн болгон уншдаг байв — тулгалт
    бүхэлдээ буруу болно.
    """
    cands: list[tuple[int, str, int]] = []
    for idx, cell in enumerate(header):
        for field, score in _score_header_cell(_cell_text(cell)).items():
            cands.append((score, field, idx))
    cands.sort(key=lambda c: (-c[0], c[2]))
    used_fields: dict[str, int] = {}
    used_cols: set[int] = set()
    for score, field, idx in cands:
        if field in used_fields or idx in used_cols:
            continue
        used_fields[field] = idx
        used_cols.add(idx)
    return used_fields


def _find_header(rows: list[list], scan: int = 25) -> tuple[int, dict[str, int]]:
    """Толгойн мөрийг эхний мөрүүдээс хайна (экспортод гарчиг/мета мөр байдаг)."""
    best_i, best_cols, best_score = -1, {}, 0
    for i, row in enumerate(rows[:scan]):
        cols = _map_columns(row)
        if "total" not in cols or not ({"date", "ddtd"} & set(cols)):
            continue
        score = len(cols) * 10 + sum(
            _score_header_cell(_cell_text(row[c])).get(f, 0) for f, c in cols.items()
        )
        if score > best_score:
            best_i, best_cols, best_score = i, cols, score
    return best_i, best_cols


# ------------------------------------------------------- утгын хөрвүүлэлт

_AMOUNT_CLEAN = re.compile(r"[^\d,.\-()]")


def _to_minor(value) -> int | None:
    """Мөнгөн дүнг мөнгө (1/100) болгоно. Decimal-аар — float бөөрөнхийлөлтгүй."""
    if value is None or value == "" or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value * 100
    if isinstance(value, (float, Decimal)):
        try:
            return int((Decimal(str(value)) * 100).to_integral_value())
        except InvalidOperation:
            return None
    s = str(value).strip()
    if not s:
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = _AMOUNT_CLEAN.sub("", s).replace("(", "").replace(")", "")
    if not s or s in ("-", ".", ","):
        return None
    # 1,320,000.00 болон 1 320 000,00 хоёуланг дэмжинэ
    if "," in s and "." in s:
        s = s.replace(",", "") if s.rfind(".") > s.rfind(",") else s.replace(".", "").replace(",", ".")
    elif s.count(",") == 1 and len(s.split(",")[-1]) in (1, 2):
        s = s.replace(",", ".")
    else:
        s = s.replace(",", "")
    try:
        d = Decimal(s)
    except InvalidOperation:
        return None
    if neg:
        d = -d
    return int((d * 100).to_integral_value())


_DATE_FORMATS = ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%d.%m.%Y", "%d/%m/%Y",
                 "%m/%d/%Y", "%Y%m%d")


def _to_iso_date(value) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, _datetime):
        return value.date().isoformat()
    if isinstance(value, _date):
        return value.isoformat()
    if isinstance(value, (int, float)) and 20000 < float(value) < 80000:
        return (_date(1899, 12, 30) + _timedelta(days=int(value))).isoformat()
    s = str(value).strip()
    if not s:
        return None
    head = s.split(" ")[0].split("T")[0]
    for fmt in _DATE_FORMATS:
        try:
            return _datetime.strptime(head, fmt).date().isoformat()
        except ValueError:
            continue
    return None


# ------------------------------------------------------- үндсэн задлагч

def parse_ebarimt_export(data: bytes, filename: str | None = None,
                         dataset: str | None = None) -> dict:
    """Нэг eBarimt экспорт файлыг задалж, тайлангийн төрөлтэй нь хамт буцаана.

    Буцаах бүтэц:
      {"dataset", "label", "direction", "party_kind", "items", "columns",
       "skipped_rows", "voided_rows", "file"}
    """
    try:
        rows = _read_rows(data, filename)
    except EbarimtParseError:
        raise
    except Exception as e:
        raise EbarimtParseError(
            f"'{filename or 'файл'}'-ыг задлаж чадсангүй ({e}). eBarimt-аас "
            f"татсан .xlsx / .xls / .csv файл эсэхийг шалгана уу.") from e
    if not rows:
        raise EbarimtParseError(
            f"'{filename or 'файл'}' хоосон эсвэл уншигдахгүй байна.")

    h_idx, cols = _find_header(rows)
    if h_idx < 0:
        seen = []
        for row in rows[:8]:
            texts = [t for t in (_cell_text(c) for c in row) if t]
            if texts:
                seen.append(" | ".join(texts[:8]))
        raise EbarimtParseError(
            f"'{filename or 'файл'}': толгойн мөр танигдсангүй. «Нийт дүн» ба "
            f"«Огноо» эсвэл «ДДТД» багана байх шаардлагатай. Уншсан эхний "
            f"мөрүүд: " + " // ".join(seen[:4]))

    header_text = " ".join(_cell_text(c) for c in rows[h_idx])
    title_text = " ".join(_cell_text(c) for r in rows[:h_idx] for c in r)
    ds = dataset or detect_dataset(filename, title_text, header_text)
    meta = DATASET_META.get(ds or "", {})

    items: list[dict] = []
    skipped = voided = 0

    for r_i, row in enumerate(rows[h_idx + 1:], start=h_idx + 2):
        def cell(field, _row=row):
            idx = cols.get(field)
            return _row[idx] if idx is not None and idx < len(_row) else None

        if not any(_cell_text(c) for c in row):
            continue

        status_text = _cell_text(cell("status")).lower()
        if status_text and any(w in status_text for w in _VOID_WORDS):
            voided += 1
            continue

        total = _to_minor(cell("total"))
        if total is None or total == 0:
            skipped += 1
            continue

        iso = _to_iso_date(cell("date"))
        ddtd = _cell_text(cell("ddtd"))
        party = _cell_text(cell("party"))
        tin = _cell_text(cell("party_tin"))

        # Нийлбэрийн («Нийт», «Дүн») мөрийг таних — огноо ч, ДДТД ч байхгүй
        if not iso and not ddtd:
            skipped += 1
            continue

        direction = meta.get("direction")
        is_return = total < 0
        if is_return and direction:
            direction = "out" if direction == "in" else "in"

        items.append({
            "date": iso,
            "total_minor": abs(total),
            "vat_minor": _to_minor(cell("vat")),
            "city_tax_minor": _to_minor(cell("city_tax")),
            "receipt_id": ddtd or f"EB-{len(items) + 1}",
            "party": party or meta.get("label") or "eBarimt",
            "party_tin": tin or None,
            "dataset": ds,
            "dataset_label": meta.get("label"),
            "direction": direction,
            "party_kind": meta.get("party_kind"),
            "is_return": is_return,
            "source_file": filename,
            "source_row": r_i,
        })

    if not items:
        raise EbarimtParseError(
            f"'{filename or 'файл'}': толгойн мөр олдсон ч дүнтэй мөр "
            f"уншигдсангүй ({skipped} мөр алгасагдав).")

    return {
        "dataset": ds,
        "label": meta.get("label"),
        "direction": meta.get("direction"),
        "party_kind": meta.get("party_kind"),
        "items": items,
        "columns": {f: int(i) for f, i in cols.items()},
        "skipped_rows": skipped,
        "voided_rows": voided,
        "file": filename,
    }


def parse_ebarimt_excel(file_bytes: bytes, filename: str | None = None) -> list[dict]:
    """Хуучин дуудлагатай нийцтэй бүрхүүл — зөвхөн мөрүүдийг буцаана."""
    return parse_ebarimt_export(file_bytes, filename)["items"]

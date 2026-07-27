"""LLM Descriptor Generator — шинэ банкны хуулгын багануудыг AI ашиглан таньж, YAML загварыг автоматаар үүсгэнэ.
"""

from __future__ import annotations

import logging
import os
import yaml
from pathlib import Path
from typing import Any

def load_excel_grid(path: Path) -> list[list[Any]]:
    try:
        from python_calamine import CalamineWorkbook
        wb = CalamineWorkbook.from_path(str(path))
        return wb.get_sheet_by_name(wb.sheet_names[0]).to_python()
    except Exception:
        pass
        
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        return [[cell.value for cell in row] for row in ws.iter_rows()]
    finally:
        wb.close()


logger = logging.getLogger(__name__)

# Сонголтоор AI сангуудыг дуудна
try:
    import google.generativeai as genai
    HAS_GEMINI = True
except ImportError:
    HAS_GEMINI = False


def _detect_columns_fallback(grid: list[list[Any]]) -> dict[str, Any]:
    """Детерминистик аргаар багануудыг таньж, YAML-д тохирох тохиргоог үүсгэнэ."""
    columns_spec = {}
    
    # Эхний 15 мөрнөөс толгой мөрийг хайна
    header_row = []
    header_idx = 0
    for r_idx, row in enumerate(grid[:15]):
        row_str = [str(cell).strip().lower() for cell in row if cell is not None]
        # Хэрэв "огноо" эсвэл "утга" орсон бол толгой мөр гэж таамаглана
        if any("огноо" in s or "date" in s or "утга" in s or "гүйлгээ" in s for s in row_str):
            header_row = [str(c).strip() for c in row if c is not None]
            header_idx = r_idx
            break
            
    if not header_row and grid:
        # Толгой мөр олдоогүй бол эхний мөрийг авна
        header_row = [f"Column {i}" for i in range(len(grid[0]))]

    # Багануудыг ухаалгаар тааруулах
    for idx, col_name in enumerate(header_row):
        col_lower = col_name.lower()
        if "огноо" in col_lower or "date" in col_lower or "хугацаа" in col_lower:
            columns_spec["date"] = {"aliases": [col_name], "type": "date", "format": "YYYY-MM-DD|YYYY/MM/DD"}
        elif "дебит" in col_lower or "debit" in col_lower or "зарлага" in col_lower:
            columns_spec["debit"] = {"aliases": [col_name], "type": "amount"}
        elif "кредит" in col_lower or "credit" in col_lower or "орлого" in col_lower:
            columns_spec["credit"] = {"aliases": [col_name], "type": "amount"}
        elif "дүн" in col_lower or "amount" in col_lower or "хэмжээ" in col_lower:
            columns_spec["amount"] = {"aliases": [col_name], "type": "amount"}
        elif "утга" in col_lower or "description" in col_lower or "тайлбар" in col_lower or "гүйлгээ" in col_lower:
            columns_spec["description"] = {"aliases": [col_name], "type": "text"}
        elif "үлдэгдэл" in col_lower or "balance" in col_lower:
            columns_spec["balance"] = {"aliases": [col_name], "type": "amount", "optional": True}
        elif "харилцагч" in col_lower or "counterparty" in col_lower or "хэнд" in col_lower:
            columns_spec["counterparty"] = {"aliases": [col_name], "type": "text", "optional": True}

    # Чухал баганууд байхгүй бол анхны утгаар тохируулна
    if "date" not in columns_spec:
        columns_spec["date"] = {"index": 0, "type": "date", "format": "YYYY-MM-DD"}
    if "description" not in columns_spec:
        columns_spec["description"] = {"index": 1, "type": "text"}
    if "amount" not in columns_spec and "debit" not in columns_spec:
        columns_spec["amount"] = {"index": 2, "type": "amount"}

    return {
        "header_row_index": header_idx,
        "columns": columns_spec
    }


def generate_yaml_descriptor(file_path: Path, filename: str, bank_name: str = "custom") -> str:
    """Excel хуулгаас YAML descriptor-ийг LLM (эсвэл fallback) ашиглан үүсгэнэ."""
    
    # 1. Excel унших
    grid = load_excel_grid(file_path)
    if not grid:
        raise ValueError("Excel файл хоосон эсвэл уншиж чадсангүй.")

    # Эхний 20 мөрийг текст болгож бэлдэнэ
    rows_preview = []
    for r in grid[:20]:
        rows_preview.append(", ".join([str(c) if c is not None else "" for c in r]))
    preview_text = "\n".join(rows_preview)

    desc_id = f"{bank_name.lower()}_{int(os.path.getmtime(file_path))}"

    # 2. Gemini AI ашиглаж үүсгэх оролдлого
    api_key = os.environ.get("GEMINI_API_KEY")
    if HAS_GEMINI and api_key:
        try:
            genai.configure(api_key=api_key)
            model = genai.GenerativeModel("gemini-2.0-flash")
            
            prompt = f"""
Та санхүүгийн SaaS системийн ухаалаг админ туслах ба банкны хуулгын багануудыг таних YAML загвар үүсгэх даалгавартай.
Дараах Excel хуулганы эхний 20 мөрийн өгөгдлийг шинжлээд, тохирох YAML descriptor загварыг зөвхөн YAML форматаар буцаана уу.

Файлын нэр: {filename}
Банкны нэр: {bank_name}

Өгөгдлийн мөрүүд:
---
{preview_text}
---

Тавигдах шаардлага:
1. Зөвхөн цэвэр YAML кодыг буцаана уу (markdown tag ```yaml хэрэггүй).
2. Загварт дараах бүтэцтэй байна:
   id: "{desc_id}"
   bank: "{bank_name}"
   file_type: "excel"
   status: "active"
   fingerprint:
     filename_pattern: ".*{bank_name}.*"
     sheet_keywords: ["хуулга", "гүйлгээ", "statement"]
     column_signature: [<огноо, дүн, эсвэл гүйлгээний утгыг илэрхийлэх баганын нэрүүд>]
     min_score: 6
   table:
     header_row_index: <толгой мөр орсон 0-ээс эхлэх индекс>
     direction_mode: "separate" (хэрэв дебит/кредит тусдаа бол) эсвэл "single" (хэрэв ганц дүнгийн багана хасах/нэмэх тэмдэгтэй бол)
     amount_locale:
       thousands: ","
       decimal: "."
     columns:
       date:
         aliases: [<огнооны баганын нэр>]
         type: "date"
         format: "YYYY-MM-DD"
       description:
         aliases: [<гүйлгээний утга, тайлбар баганын нэр>]
         type: "text"
       debit: (direction_mode separate үед)
         aliases: [<дебит, зарлага баганын нэр>]
         type: "amount"
       credit: (direction_mode separate үед)
         aliases: [<кредит, орлого баганын нэр>]
         type: "amount"
       amount: (direction_mode single үед)
         aliases: [<дүнгийн баганын нэр>]
         type: "amount"
"""
            res = model.generate_content(prompt)
            yaml_text = res.text.strip()
            # ```yaml ... ``` хавчуургыг арилгах
            if yaml_text.startswith("```"):
                lines = yaml_text.splitlines()
                if lines[0].startswith("```"):
                    lines = lines[1:]
                if lines[-1].startswith("```"):
                    lines = lines[:-1]
                yaml_text = "\n".join(lines).strip()
            
            # YAML формат зөв эсэхийг баталгаажуулна
            yaml.safe_load(yaml_text)
            return yaml_text
        except Exception as e:
            logger.warning("Gemini AI descriptor generation failed, using fallback: %s", e)

    # 3. Fallback: Детерминистик код
    fallback_data = _detect_columns_fallback(grid)
    
    direction_mode = "separate" if "debit" in fallback_data["columns"] else "single"
    
    descriptor_dict = {
        "id": desc_id,
        "bank": bank_name,
        "file_type": "excel",
        "status": "active",
        "fingerprint": {
            "filename_pattern": f".*{bank_name.lower()}.*",
            "sheet_keywords": ["хуулга", "гүйлгээ", "statement"],
            "column_signature": [v["aliases"][0] for k, v in fallback_data["columns"].items() if "aliases" in v],
            "min_score": 6
        },
        "table": {
            "header_row_index": fallback_data["header_row_index"],
            "direction_mode": direction_mode,
            "amount_locale": {
                "thousands": ",",
                "decimal": "."
            },
            "columns": fallback_data["columns"]
        },
        "validation": {
            "continuity": True
        }
    }
    
    return yaml.dump(descriptor_dict, allow_unicode=True, sort_keys=False)

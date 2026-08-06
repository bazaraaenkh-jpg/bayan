"""СТ-1..4 санхүүгийн тайлангуудыг e-balance.mn / e-tax.mn загвараар XML болон Excel рүү экспортлох модуль.
"""

from __future__ import annotations

import tempfile
import xml.etree.ElementTree as ET
from datetime import date
from pathlib import Path
from xml.dom import minidom

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import reports
from .models import Company


def _minor_to_unit(minor_val: int | None) -> float:
    """Minor unit (мөнгө)-ийг үндсэн нэгж (төгрөг) рүү хөрвүүлнэ."""
    if minor_val is None:
        return 0.0
    return float(minor_val) / 100.0


# ================================================================= XML ЭКСПОРТ (e-balance.mn)

def export_xml(session, company_id: str, year: int, month: int) -> str:
    """Санхүүгийн 4 тайланг нэгтгэсэн e-balance.mn-тэй нийцтэй XML бүтцийг үүсгэнэ."""
    company = session.get(Company, company_id)
    company_name = company.name if company else "Тодорхойгүй"
    reg_no = company.reg_no if company else ""

    # Огнооны мужуудыг бэлдэнэ
    date_from = date(year, 1, 1)
    date_to = date(year, month, 28)  # Сарын эхний ~28 хоногоор (or last day)
    # Сарын сүүлийн өдрийг олно
    import calendar
    last_day = calendar.monthrange(year, month)[1]
    date_to = date(year, month, last_day)

    # Тайлангуудыг бодно
    bs = reports.balance_sheet(session, company_id, date_to)
    inc = reports.income_statement(session, company_id, date_from, date_to)
    eq = reports.statement_of_changes_in_equity(session, company_id, date_from, date_to)
    cf = reports.cash_flow_statement(session, company_id, date_from, date_to)

    # XML мод үүсгэх
    root = ET.Element("eBalance")
    
    # 1. Толгой хэсэг
    header = ET.SubElement(root, "Header")
    ET.SubElement(header, "CompanyRegister").text = reg_no
    ET.SubElement(header, "CompanyName").text = company_name
    ET.SubElement(header, "Year").text = str(year)
    ET.SubElement(header, "Month").text = str(month)
    ET.SubElement(header, "ExportDate").text = date.today().isoformat()

    # 2. СТ-1: Баланс (Balance Sheet)
    st1 = ET.SubElement(root, "BalanceSheet")
    for group_name, items in [("Assets", bs["assets"]), ("Liabilities", bs["liabilities"]), ("Equity", bs["equity"])]:
        group_el = ET.SubElement(st1, group_name)
        for item in items:
            row_el = ET.SubElement(group_el, "Row")
            ET.SubElement(row_el, "Name").text = item["name"]
            ET.SubElement(row_el, "Amount").text = f"{_minor_to_unit(item['amount_minor']):.2f}"
            ET.SubElement(row_el, "PriorAmount").text = f"{_minor_to_unit(item['prior_amount_minor']):.2f}"

    # 3. СТ-2: Орлогын тайлан (Income Statement)
    st2 = ET.SubElement(root, "IncomeStatement")
    for key, name in [
        ("revenue_minor", "Борлуулалтын орлого"),
        ("cogs_minor", "Борлуулсан бүтээгдэхүүний өртөг (ББӨ)"),
        ("gross_profit_minor", "Нийт ашиг (алдагдал)"),
        ("expenses_minor", "Үйл ажиллагааны зардал"),
        ("net_income_minor", "Тайлант үеийн цэвэр ашиг (алдагдал)"),
    ]:
        row_el = ET.SubElement(st2, "Row")
        ET.SubElement(row_el, "Name").text = name
        ET.SubElement(row_el, "Amount").text = f"{_minor_to_unit(inc.get(key)):.2f}"
        ET.SubElement(row_el, "PriorAmount").text = f"{_minor_to_unit(inc.get('prior_' + key)):.2f}"

    # 4. СТ-3: Өмчийн өөрчлөлтийн тайлан (Statement of Changes in Equity)
    st3 = ET.SubElement(root, "ChangesInEquity")
    curr_eq = eq["current"]
    prior_eq = eq.get("prior", {})
    
    for state_key, name in [
        ("start", "Эхний үлдэгдэл"),
        ("net_income", "Тайлант үеийн цэвэр ашиг"),
        ("changes", "Өөрчлөлт"),
        ("end", "Эцсийн үлдэгдэл"),
    ]:
        row_el = ET.SubElement(st3, "Row")
        ET.SubElement(row_el, "Type").text = name
        
        # Current year
        c_state = curr_eq.get(state_key, {})
        ET.SubElement(row_el, "Capital").text = f"{_minor_to_unit(c_state.get('capital')):.2f}"
        ET.SubElement(row_el, "Retained").text = f"{_minor_to_unit(c_state.get('retained')):.2f}"
        ET.SubElement(row_el, "Total").text = f"{_minor_to_unit(c_state.get('total')):.2f}"
        
        # Prior year
        p_state = prior_eq.get(state_key, {})
        ET.SubElement(row_el, "PriorCapital").text = f"{_minor_to_unit(p_state.get('capital')):.2f}"
        ET.SubElement(row_el, "PriorRetained").text = f"{_minor_to_unit(p_state.get('retained')):.2f}"
        ET.SubElement(row_el, "PriorTotal").text = f"{_minor_to_unit(p_state.get('total')):.2f}"

    # 5. СТ-4: Мөнгөн гүйлгээний тайлан (Statement of Cash Flows)
    st4 = ET.SubElement(root, "CashFlow")
    curr_cf = cf["current"]
    prior_cf = cf.get("prior", {})
    
    for key, name in [
        ("operating_inflow", "Үйл ажиллагааны мөнгөн орлого"),
        ("operating_outflow", "Үйл ажиллагааны мөнгөн зарлага"),
        ("operating_net", "Үйл ажиллагааны цэвэр мөнгөн урсгал"),
        ("investing_inflow", "Хөрөнгө оруулалтын мөнгөн орлого"),
        ("investing_outflow", "Хөрөнгө оруулалтын мөнгөн зарлага"),
        ("investing_net", "Хөрөнгө оруулалтын цэвэр мөнгөн урсгал"),
        ("financing_inflow", "Санхүүжилтийн мөнгөн орлого"),
        ("financing_outflow", "Санхүүжилтийн мөнгөн зарлага"),
        ("financing_net", "Санхүүжилтийн цэвэр мөнгөн урсгал"),
        ("net_change", "Мөнгөний цэвэр өөрчлөлт"),
    ]:
        row_el = ET.SubElement(st4, "Row")
        ET.SubElement(row_el, "Name").text = name
        ET.SubElement(row_el, "Amount").text = f"{_minor_to_unit(curr_cf.get(key)):.2f}"
        ET.SubElement(row_el, "PriorAmount").text = f"{_minor_to_unit(prior_cf.get(key)):.2f}"

    # XML-ийг гоёмсог форматлах (pretty-print)
    xml_str = ET.tostring(root, encoding="utf-8")
    parsed = minidom.parseString(xml_str)
    return parsed.toprettyxml(indent="  ")


# ================================================================= EXCEL ЭКСПОРТ (Openpyxl)

def export_excel(session, company_id: str, year: int, month: int) -> Path:
    """СТ-1..4 санхүүгийн 4 тайланг агуулсан загвар бүхий Excel файл үүсгэнэ."""
    company = session.get(Company, company_id)
    company_name = company.name if company else "Тодорхойгүй"
    
    import calendar
    last_day = calendar.monthrange(year, month)[1]
    date_from = date(year, 1, 1)
    date_to = date(year, month, last_day)

    bs = reports.balance_sheet(session, company_id, date_to)
    inc = reports.income_statement(session, company_id, date_from, date_to)
    eq = reports.statement_of_changes_in_equity(session, company_id, date_from, date_to)
    cf = reports.cash_flow_statement(session, company_id, date_from, date_to)

    wb = Workbook()
    
    # Загварууд (Styling)
    font_title = Font(name="Calibri", size=14, bold=True)
    font_subtitle = Font(name="Calibri", size=10, italic=True)
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_normal = Font(name="Calibri", size=11)
    
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Sleek Navy Blue
    fill_total = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Light Gray for totals
    
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    border_thin = Border(bottom=Side(style='thin', color="D9D9D9"))
    border_double = Border(top=Side(style='thin', color="000000"), bottom=Side(style='double', color="000000"))

    # Sheet 1: СТ-1 (Санхүү байдлын тайлан)
    ws1 = wb.active
    ws1.title = "СТ-1 Баланс"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.append([company_name])
    ws1.append(["СТ-1: Санхүү байдлын тайлан"])
    ws1.append([f"Тайлант хугацаа: {year} оны {month}-р сарын {last_day}-ний байдлаар"])
    ws1.append([f"Хэмжих нэгж: Төгрөг (₮)"])
    ws1.append([]) # хоосон мөр
    
    ws1.cell(row=1, column=1).font = font_title
    ws1.cell(row=2, column=1).font = font_bold
    ws1.cell(row=3, column=1).font = font_subtitle
    ws1.cell(row=4, column=1).font = font_subtitle

    # Баганы гарчиг
    headers1 = ["Үзүүлэлт", "Тайлант оны үлдэгдэл", "Өмнөх оны үлдэгдэл"]
    ws1.append(headers1)
    for col_idx in range(1, 4):
        cell = ws1.cell(row=6, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx > 1 else align_left
        
    # Assets
    ws1.append(["АКТИВ ХӨРӨНГӨ"])
    ws1.cell(row=7, column=1).font = font_bold
    for item in bs["assets"]:
        ws1.append([item["name"], _minor_to_unit(item["amount_minor"]), _minor_to_unit(item["prior_amount_minor"])])
    
    # Asset Total
    asset_total_row = ws1.max_row + 1
    ws1.append(["Нийт идэвхтэй хөрөнгө", _minor_to_unit(bs["total_assets_minor"]), _minor_to_unit(bs["prior_total_assets_minor"])])
    for col_idx in range(1, 4):
        c = ws1.cell(row=asset_total_row, column=col_idx)
        c.font = font_bold
        c.fill = fill_total
        c.border = border_double
        
    # Liabilities & Equity
    ws1.append([])
    ws1.append(["ӨР ТӨЛБӨР БА ЭЗДИЙН ӨМЧ"])
    ws1.cell(row=asset_total_row + 2, column=1).font = font_bold
    
    for item in bs["liabilities"]:
        ws1.append([item["name"], _minor_to_unit(item["amount_minor"]), _minor_to_unit(item["prior_amount_minor"])])
        
    for item in bs["equity"]:
        ws1.append([item["name"], _minor_to_unit(item["amount_minor"]), _minor_to_unit(item["prior_amount_minor"])])

    # Total LE
    le_total_row = ws1.max_row + 1
    ws1.append(["Нийт өр төлбөр ба эздийн өмч", _minor_to_unit(bs["total_liab_equity_minor"]), _minor_to_unit(bs["prior_total_liab_equity_minor"])])
    for col_idx in range(1, 4):
        c = ws1.cell(row=le_total_row, column=col_idx)
        c.font = font_bold
        c.fill = fill_total
        c.border = border_double

    # Формат тохируулах
    for r in range(8, ws1.max_row + 1):
        if r not in (asset_total_row, le_total_row, asset_total_row + 1, asset_total_row + 2):
            ws1.cell(row=r, column=1).font = font_normal
            ws1.cell(row=r, column=1).border = border_thin
            for col_idx in (2, 3):
                c = ws1.cell(row=r, column=col_idx)
                c.number_format = "#,##0.00"
                c.font = font_normal
                c.alignment = align_right
                c.border = border_thin

    # Column widths
    ws1.column_dimensions['A'].width = 38
    ws1.column_dimensions['B'].width = 24
    ws1.column_dimensions['C'].width = 24


    # Sheet 2: СТ-2 (Орлогын тайлан)
    ws2 = wb.create_sheet(title="СТ-2 Орлого")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.append([company_name])
    ws2.append(["СТ-2: Үйл ажиллагааны орлогын тайлан"])
    ws2.append([f"Тайлант хугацаа: {year} оны {month}-р сарын {last_day} хүртэл"])
    ws2.append([f"Хэмжих нэгж: Төгрөг (₮)"])
    ws2.append([])
    
    ws2.cell(row=1, column=1).font = font_title
    ws2.cell(row=2, column=1).font = font_bold
    ws2.cell(row=3, column=1).font = font_subtitle
    ws2.cell(row=4, column=1).font = font_subtitle

    headers2 = ["Үзүүлэлт", "Тайлант оны дүн", "Өмнөх оны дүн"]
    ws2.append(headers2)
    for col_idx in range(1, 4):
        cell = ws2.cell(row=6, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx > 1 else align_left

    rows_st2 = [
        ("Үйл ажиллагааны борлуулалтын орлого", _minor_to_unit(inc["revenue_minor"]), _minor_to_unit(inc.get("prior_revenue_minor"))),
        ("Борлуулсан бүтээгдэхүүний өртөг (ББӨ)", _minor_to_unit(inc["cogs_minor"]), _minor_to_unit(inc.get("prior_cogs_minor"))),
        ("Нийт үйл ажиллагааны ашиг (алдагдал)", _minor_to_unit(inc["gross_profit_minor"]), _minor_to_unit(inc.get("prior_gross_profit_minor"))),
        ("Үйл ажиллагааны зардал (нийт)", _minor_to_unit(inc["expenses_minor"]), _minor_to_unit(inc.get("prior_expenses_minor"))),
        ("Үйл ажиллагааны цэвэр ашиг (алдагдал)", _minor_to_unit(inc["net_income_minor"]), _minor_to_unit(inc.get("prior_net_income_minor"))),
    ]

    for idx, r in enumerate(rows_st2, start=7):
        ws2.append([r[0], r[1], r[2]])
        # Нийт ашиг болон Цэвэр ашиг мөрүүдийг тодруулна
        is_bold_row = idx in (9, 11)
        for col_idx in range(1, 4):
            c = ws2.cell(row=idx, column=col_idx)
            c.font = font_bold if is_bold_row else font_normal
            if is_bold_row:
                c.fill = fill_total
                c.border = border_double
            else:
                c.border = border_thin
            if col_idx > 1:
                c.number_format = "#,##0.00"
                c.alignment = align_right

    ws2.column_dimensions['A'].width = 38
    ws2.column_dimensions['B'].width = 24
    ws2.column_dimensions['C'].width = 24


    # Sheet 3: СТ-3 (Өмчийн өөрчлөлт)
    ws3 = wb.create_sheet(title="СТ-3 Өмч")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.append([company_name])
    ws3.append(["СТ-3: Эздийн өмчийн өөрчлөлтийн тайлан"])
    ws3.append([f"Тайлант хугацаа: {year} оны {month}-р сарын {last_day} хүртэл"])
    ws3.append([])
    
    ws3.cell(row=1, column=1).font = font_title
    ws3.cell(row=2, column=1).font = font_bold
    ws3.cell(row=3, column=1).font = font_subtitle

    headers3 = ["Үзүүлэлт", "Хувь нийлүүлсэн хөрөнгө", "Хуримтлагдсан ашиг", "Нийт өмч"]
    ws3.append(headers3)
    for col_idx in range(1, 5):
        cell = ws3.cell(row=5, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx > 1 else align_left

    curr_eq = eq["current"]
    rows_st3 = [
        ("Эхний үлдэгдэл", _minor_to_unit(curr_eq["start"]["capital"]), _minor_to_unit(curr_eq["start"]["retained"]), _minor_to_unit(curr_eq["start"]["total"])),
        ("Тайлант үеийн цэвэр ашиг", _minor_to_unit(curr_eq["net_income"]["capital"]), _minor_to_unit(curr_eq["net_income"]["retained"]), _minor_to_unit(curr_eq["net_income"]["total"])),
        ("Бусад өөрчлөлт / Ногдол ашиг", _minor_to_unit(curr_eq["changes"]["capital"]), _minor_to_unit(curr_eq["changes"]["retained"]), _minor_to_unit(curr_eq["changes"]["total"])),
        ("Эцсийн үлдэгдэл", _minor_to_unit(curr_eq["end"]["capital"]), _minor_to_unit(curr_eq["end"]["retained"]), _minor_to_unit(curr_eq["end"]["total"])),
    ]

    for idx, r in enumerate(rows_st3, start=6):
        ws3.append([r[0], r[1], r[2], r[3]])
        is_bold_row = idx in (6, 9)
        for col_idx in range(1, 5):
            c = ws3.cell(row=idx, column=col_idx)
            c.font = font_bold if is_bold_row else font_normal
            if is_bold_row:
                c.fill = fill_total
                c.border = border_double
            else:
                c.border = border_thin
            if col_idx > 1:
                c.number_format = "#,##0.00"
                c.alignment = align_right

    ws3.column_dimensions['A'].width = 28
    ws3.column_dimensions['B'].width = 24
    ws3.column_dimensions['C'].width = 24
    ws3.column_dimensions['D'].width = 24


    # Sheet 4: СТ-4 (Мөнгөн гүйлгээ)
    ws4 = wb.create_sheet(title="СТ-4 Мөнгөн гүйлгээ")
    ws4.views.sheetView[0].showGridLines = True
    
    ws4.append([company_name])
    ws4.append(["СТ-4: Мөнгөн гүйлгээний тайлан (Шууд арга)"])
    ws4.append([f"Тайлант хугацаа: {year} оны {month}-р сарын {last_day} хүртэл"])
    ws4.append([f"Хэмжих нэгж: Төгрөг (₮)"])
    ws4.append([])
    
    ws4.cell(row=1, column=1).font = font_title
    ws4.cell(row=2, column=1).font = font_bold
    ws4.cell(row=3, column=1).font = font_subtitle
    ws4.cell(row=4, column=1).font = font_subtitle

    headers4 = ["Мөнгөн гүйлгээний үзүүлэлтүүд", "Тайлант оны дүн", "Өмнөх оны дүн"]
    ws4.append(headers4)
    for col_idx in range(1, 4):
        cell = ws4.cell(row=6, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx > 1 else align_left

    curr_cf = cf["current"]
    prior_cf = cf.get("prior", {})
    
    rows_st4 = [
        ("Үйл ажиллагааны мөнгөн орлого", _minor_to_unit(curr_cf["operating_inflow"]), _minor_to_unit(prior_cf.get("operating_inflow"))),
        ("Үйл ажиллагааны мөнгөн зарлага", _minor_to_unit(curr_cf["operating_outflow"]), _minor_to_unit(prior_cf.get("operating_outflow"))),
        ("Үйл ажиллагааны цэвэр мөнгөн урсгал", _minor_to_unit(curr_cf["operating_net"]), _minor_to_unit(prior_cf.get("operating_net"))),
        ("Хөрөнгө оруулалтын мөнгөн орлого", _minor_to_unit(curr_cf["investing_inflow"]), _minor_to_unit(prior_cf.get("investing_inflow"))),
        ("Хөрөнгө оруулалтын мөнгөн зарлага", _minor_to_unit(curr_cf["investing_outflow"]), _minor_to_unit(prior_cf.get("investing_outflow"))),
        ("Хөрөнгө оруулалтын цэвэр мөнгөн урсгал", _minor_to_unit(curr_cf["investing_net"]), _minor_to_unit(prior_cf.get("investing_net"))),
        ("Санхүүжилтийн мөнгөн орлого", _minor_to_unit(curr_cf["financing_inflow"]), _minor_to_unit(prior_cf.get("financing_inflow"))),
        ("Санхүүжилтийн мөнгөн зарлага", _minor_to_unit(curr_cf["financing_outflow"]), _minor_to_unit(prior_cf.get("financing_outflow"))),
        ("Санхүүжилтийн цэвэр мөнгөн урсгал", _minor_to_unit(curr_cf["financing_net"]), _minor_to_unit(prior_cf.get("financing_net"))),
        ("Мөнгө, түүнтэй адилтгах хөрөнгийн цэвэр өөрчлөлт", _minor_to_unit(curr_cf["net_change"]), _minor_to_unit(prior_cf.get("net_change"))),
    ]

    for idx, r in enumerate(rows_st4, start=7):
        ws4.append([r[0], r[1], r[2]])
        is_bold_row = idx in (9, 12, 15, 16)
        for col_idx in range(1, 4):
            c = ws4.cell(row=idx, column=col_idx)
            c.font = font_bold if is_bold_row else font_normal
            if is_bold_row:
                c.fill = fill_total
                c.border = border_double
            else:
                c.border = border_thin
            if col_idx > 1:
                c.number_format = "#,##0.00"
                c.alignment = align_right

    ws4.column_dimensions['A'].width = 46
    ws4.column_dimensions['B'].width = 24
    ws4.column_dimensions['C'].width = 24

    # ---------------------------------------------------------------- Хуудас 5-9
    # СТ-1..4 нь нэгтгэсэн дүн харуулдаг. Нягтлан бодогчид дүн бүрийн ард
    # ямар данс, харилцагч, бараа байгааг харах шаардлагатай тул дэлгэрэнгүй
    # хуудсуудыг нэг файлд хамт гаргана.
    style = _SheetStyle(font_title, font_header, font_bold, fill_header,
                        fill_total, align_left, align_right, align_center)

    _sheet_trial_balance(wb, style, session, company_id, company_name, date_from, date_to)
    _sheet_aging(wb, style, session, company_id, company_name, date_to)
    _sheet_inventory(wb, style, session, company_id, company_name, date_to)
    _sheet_fixed_assets(wb, style, session, company_id, company_name, date_to)
    _sheet_vat(wb, style, session, company_id, company_name, year, month)

    # Түр файл үүсгэж хадгална
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="bayan_report_")
    tmp_path = Path(tmp.name)
    tmp.close()

    wb.save(tmp_path)
    return tmp_path


# ==================================================== Дэлгэрэнгүй хуудсууд

class _SheetStyle:
    """export_excel-ийн загваруудыг дэлгэрэнгүй хуудсууд руу дамжуулна."""

    def __init__(self, title, header, bold, fill_header, fill_total,
                 left, right, center):
        self.title, self.header, self.bold = title, header, bold
        self.fill_header, self.fill_total = fill_header, fill_total
        self.left, self.right, self.center = left, right, center


def _new_sheet(wb, style, title: str, company_name: str, subtitle: str,
               headers: list[tuple[str, int]]):
    """Толгой мэдээлэл, баганын нэр бүхий хуудас бэлдэнэ."""
    ws = wb.create_sheet(title=title)
    ws["A1"] = company_name
    ws["A1"].font = style.title
    ws["A2"] = subtitle
    ws["A2"].font = style.bold
    ws.append([])

    row = 4
    for idx, (name, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx, value=name)
        cell.font = style.header
        cell.fill = style.fill_header
        cell.alignment = style.center
        ws.column_dimensions[cell.column_letter].width = width
    # Мөрийн хаягаар царцаана — ws.cell(...) дуудвал хоосон нүд үүсч, дараагийн
    # append нэг мөр доошилно
    ws.freeze_panes = f"A{row + 1}"
    return ws


def _write_rows(ws, style, rows: list[list], num_from: int = 2,
                total_row: list | None = None, empty_note: str = "Мэдээлэл алга."):
    """Мөрүүдийг бичиж, тоон баганад формат тавина."""
    if not rows:
        ws.append([empty_note])
        return
    for r in rows:
        ws.append(r)
        for col in range(num_from, len(r) + 1):
            c = ws.cell(row=ws.max_row, column=col)
            if isinstance(c.value, (int, float)):
                c.number_format = "#,##0.00"
                c.alignment = style.right
    if total_row:
        ws.append(total_row)
        for col in range(1, len(total_row) + 1):
            c = ws.cell(row=ws.max_row, column=col)
            c.font = style.bold
            c.fill = style.fill_total
            if isinstance(c.value, (int, float)):
                c.number_format = "#,##0.00"
                c.alignment = style.right


def _sheet_trial_balance(wb, style, session, company_id, company_name, d_from, d_to):
    """Хуудас 5 — Гүйлгээний баланс. СТ-1/СТ-2-ын дүн ямар данснаас гарсныг харуулна."""
    from .ledger import trial_balance

    ws = _new_sheet(wb, style, "5. Гүйлгээ баланс", company_name,
                    f"Гүйлгээний баланс · {d_from} — {d_to}",
                    [("Код", 12), ("Дансны нэр", 42), ("Дебит", 18),
                     ("Кредит", 18), ("Үлдэгдэл", 18)])
    tb = [r for r in trial_balance(session, company_id, d_from, d_to)
          if r["debit_minor"] or r["credit_minor"]]
    rows = [[r["code"], r["name"], _minor_to_unit(r["debit_minor"]),
             _minor_to_unit(r["credit_minor"]), _minor_to_unit(r["balance_minor"])]
            for r in tb]
    _write_rows(ws, style, rows, num_from=3, total_row=[
        "", "НИЙТ",
        _minor_to_unit(sum(r["debit_minor"] for r in tb)),
        _minor_to_unit(sum(r["credit_minor"] for r in tb)), ""],
        empty_note="Тайлант хугацаанд гүйлгээ алга.")


def _sheet_aging(wb, style, session, company_id, company_name, as_of):
    """Хуудас 6 — Авлага, өглөгийн насжилт."""
    from .partners import InvoiceKind, aging_report

    ws = _new_sheet(wb, style, "6. Авлага-Өглөг", company_name,
                    f"Харилцагчийн насжилт · {as_of}",
                    [("Төрөл", 14), ("Харилцагч", 34), ("Нэхэмжлэх", 16),
                     ("Хугацаа", 14), ("Хоцорсон хоног", 16), ("Үлдэгдэл", 18)])
    rows = []
    for kind, label in ((InvoiceKind.sales, "Авлага"), (InvoiceKind.purchase, "Өглөг")):
        for r in aging_report(session, company_id, kind, as_of):
            rows.append([label, r.get("counterparty", ""), r.get("number", ""),
                         str(r.get("due_date", "")), r.get("overdue_days", 0),
                         _minor_to_unit(r.get("outstanding_minor", 0))])
    _write_rows(ws, style, rows, num_from=5,
                empty_note="Нээлттэй авлага, өглөг алга.")


def _sheet_inventory(wb, style, session, company_id, company_name, as_of):
    """Хуудас 7 — Бараа материалын үлдэгдэл."""
    from .inventory import stock_report

    ws = _new_sheet(wb, style, "7. Бараа материал", company_name,
                    f"Барааны үлдэгдэл · {as_of}",
                    [("Код", 16), ("Барааны нэр", 38), ("Нэгж", 10),
                     ("Тоо ширхэг", 14), ("Дундаж өртөг", 16), ("Нийт өртөг", 18)])
    items = stock_report(session, company_id)
    rows = [[i.get("code", ""), i.get("name", ""), i.get("unit", ""),
             i.get("qty", 0), _minor_to_unit(i.get("avg_cost_minor", 0)),
             _minor_to_unit(i.get("total_cost_minor", 0))] for i in items]
    _write_rows(ws, style, rows, num_from=4, total_row=[
        "", "НИЙТ", "", "", "",
        _minor_to_unit(sum(i.get("total_cost_minor", 0) for i in items))],
        empty_note="Бараа материалын үлдэгдэл алга.")


def _sheet_fixed_assets(wb, style, session, company_id, company_name, as_of):
    """Хуудас 8 — Үндсэн хөрөнгө ба хуримтлагдсан элэгдэл."""
    from .assets import asset_register

    ws = _new_sheet(wb, style, "8. Үндсэн хөрөнгө", company_name,
                    f"Хөрөнгийн бүртгэл ба элэгдэл · {as_of}",
                    [("Код", 16), ("Хөрөнгийн нэр", 36), ("Ашиглах хугацаа (сар)", 20),
                     ("Анхны өртөг", 18), ("Хуримтлагдсан элэгдэл", 22),
                     ("Үлдэгдэл өртөг", 18), ("Төлөв", 12)])
    regs = asset_register(session, company_id)
    rows = [[a.get("code", ""), a.get("name", ""), a.get("months", 0),
             _minor_to_unit(a.get("cost_minor", 0)),
             _minor_to_unit(a.get("accumulated_minor", 0)),
             _minor_to_unit(a.get("book_value_minor", 0)),
             "Ашиглаж буй" if a.get("active") else "Актлагдсан"] for a in regs]
    _write_rows(ws, style, rows, num_from=3, total_row=[
        "", "НИЙТ", "",
        _minor_to_unit(sum(a.get("cost_minor", 0) for a in regs)),
        _minor_to_unit(sum(a.get("accumulated_minor", 0) for a in regs)),
        _minor_to_unit(sum(a.get("book_value_minor", 0) for a in regs)), ""],
        empty_note="Бүртгэлтэй үндсэн хөрөнгө алга.")


# ТТ-03А маягтын мөрийн дугаар ба албан ёсны нэр
_TT03A_LABELS = [
    ("1_niit_borluulalt",         "1. Нийт борлуулалт"),
    ("3_noat_nogdoh_borluulalt",  "3. НӨАТ ногдох борлуулалт"),
    ("26_nogduulsan_tatvar",      "26. Ногдуулсан татвар"),
    ("31_nogduulsan_niit",        "31. Ногдуулсан татварын нийт дүн"),
    ("32_niit_hudaldan_avalt",    "32. Нийт худалдан авалт"),
    ("33_noattai_hudaldan_avalt", "33. НӨАТ-тай худалдан авалт"),
    ("42_tolson_noat",            "42. Төлсөн НӨАТ"),
    ("43_hasagdahgui_noat",       "43. Хасагдахгүй НӨАТ"),
    ("49_hasagdah_noat",          "49. Хасагдах НӨАТ"),
    ("56_tolboh_zohih",           "56. Төлбөл зохих"),
    ("57_butsaan_avah",           "57. Буцаан авах"),
    ("64_etssiin_tolboh",         "64. Эцсийн төлбөл зохих"),
    ("65_etssiin_butsaan_avah",   "65. Эцсийн буцаан авах"),
]


def _sheet_vat(wb, style, session, company_id, company_name, year, month):
    """Хуудас 9 — НӨАТ-ын тулгалт (ТТ-03А)."""
    from .vat import tt03a

    data = tt03a(session, company_id, year, month)
    ws = _new_sheet(wb, style, "9. НӨАТ (ТТ-03А)", company_name,
                    f"НӨАТ-ын тайлан · {data.get('period', '')}",
                    [("Маягтын мөр", 46), ("Дүн (₮)", 22)])
    vals = data.get("rows", {})
    rows = [[label, _minor_to_unit(vals.get(key, 0))] for key, label in _TT03A_LABELS]
    _write_rows(ws, style, rows, num_from=2,
                empty_note="НӨАТ-ын мэдээлэл алга.")


def export_single_excel(session, company_id: str, report_type: str, date_from_str: str | None, date_to_str: str | None) -> Path:
    """Тусгай нэг тайланг (СТ-1..4, Бараа, Насжилт, НӨАТ) Excel рүү экспортолно."""
    d_from = date.fromisoformat(date_from_str) if date_from_str else None
    d_to = date.fromisoformat(date_to_str) if date_to_str else None
    
    company = session.get(Company, company_id)
    company_name = company.name if company else "Тодорхойгүй"
    
    wb = Workbook()
    ws = wb.active
    ws.views.sheetView[0].showGridLines = True
    
    font_title = Font(name="Calibri", size=14, bold=True)
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_normal = Font(name="Calibri", size=11)
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_total = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    border_thin = Border(bottom=Side(style='thin', color="D9D9D9"))
    border_double = Border(top=Side(style='thin', color="000000"), bottom=Side(style='double', color="000000"))

    if report_type == "st1":
        ws.title = "СТ-1 Баланс"
        ws.append([company_name])
        ws.append(["СТ-1: Санхүү байдлын тайлан"])
        ws.append([f"Тайлант хугацаа: {date_to_str} байдлаар"])
        ws.append(["Хэмжих нэгж: Төгрөг (₮)"])
        ws.append([])
        
        ws.cell(row=1, column=1).font = font_title
        ws.cell(row=2, column=1).font = font_bold
        
        headers = ["Үзүүлэлт", "Тайлант оны үлдэгдэл", "Өмнөх оны үлдэгдэл"]
        ws.append(headers)
        for col_idx in range(1, 4):
            cell = ws.cell(row=6, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            
        bs = reports.balance_sheet(session, company_id, d_to)
        
        def add_items(title, items):
            ws.append([title, "", ""])
            row_idx = ws.max_row
            ws.cell(row=row_idx, column=1).font = font_bold
            for item in items:
                ws.append([
                    item["name"], 
                    _minor_to_unit(item["amount_minor"]), 
                    _minor_to_unit(item["prior_amount_minor"])
                ])
                r = ws.max_row
                ws.cell(row=r, column=1).alignment = align_left
                ws.cell(row=r, column=2).alignment = align_right
                ws.cell(row=r, column=3).alignment = align_right
                ws.cell(row=r, column=2).number_format = '#,##0.00'
                ws.cell(row=r, column=3).number_format = '#,##0.00'
                
        add_items("ХӨРӨНГӨ", bs["assets"])
        ws.append(["Нийт хөрөнгө", _minor_to_unit(bs["total_assets_minor"]), _minor_to_unit(bs["prior_total_assets_minor"])])
        r = ws.max_row
        for col in range(1, 4):
            c = ws.cell(row=r, column=col)
            c.font = font_bold
            c.fill = fill_total
            c.border = border_double
        ws.cell(row=r, column=2).number_format = '#,##0.00'
        ws.cell(row=r, column=3).number_format = '#,##0.00'
        
        add_items("ӨР ТӨЛБӨР БА ЭЗДИЙН ӨМЧ", bs["liabilities"] + bs["equity"])
        ws.append(["Нийт өр төлбөр ба эздийн өмч", _minor_to_unit(bs["total_liab_equity_minor"]), _minor_to_unit(bs["prior_total_liab_equity_minor"])])
        r = ws.max_row
        for col in range(1, 4):
            c = ws.cell(row=r, column=col)
            c.font = font_bold
            c.fill = fill_total
            c.border = border_double
        ws.cell(row=r, column=2).number_format = '#,##0.00'
        ws.cell(row=r, column=3).number_format = '#,##0.00'
        
    elif report_type == "st2":
        ws.title = "СТ-2 Орлого"
        ws.append([company_name])
        ws.append(["СТ-2: Орлогын тайлан"])
        ws.append([f"Тайлант хугацаа: {date_from_str} - {date_to_str}"])
        ws.append(["Хэмжих нэгж: Төгрөг (₮)"])
        ws.append([])
        
        ws.cell(row=1, column=1).font = font_title
        ws.cell(row=2, column=1).font = font_bold
        
        headers = ["Үзүүлэлт", "Тайлант үеийн дүн", "Өмнөх оны дүн"]
        ws.append(headers)
        for col_idx in range(1, 4):
            cell = ws.cell(row=6, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            
        inc = reports.income_statement(session, company_id, d_from, d_to)
        rows_to_add = [
            ("Борлуулалтын орлого", inc["revenue_minor"], inc.get("prior_revenue_minor", 0)),
            ("Борлуулсан бүтээгдэхүүний өртөг (ББӨ)", inc["cogs_minor"], inc.get("prior_cogs_minor", 0)),
            ("Нийт ашиг (алдагдал)", inc["gross_profit_minor"], inc.get("prior_gross_profit_minor", 0)),
            ("Үйл ажиллагааны зардал", inc["expenses_minor"], inc.get("prior_expenses_minor", 0)),
            ("Цэвэр ашиг (алдагдал)", inc["net_income_minor"], inc.get("prior_net_income_minor", 0)),
        ]
        
        for name, curr, prior in rows_to_add:
            ws.append([name, _minor_to_unit(curr), _minor_to_unit(prior)])
            r = ws.max_row
            ws.cell(row=r, column=1).alignment = align_left
            ws.cell(row=r, column=2).alignment = align_right
            ws.cell(row=r, column=3).alignment = align_right
            ws.cell(row=r, column=2).number_format = '#,##0.00'
            ws.cell(row=r, column=3).number_format = '#,##0.00'
            if "ашиг" in name or "Нийт" in name or "Цэвэр" in name:
                for col in range(1, 4):
                    ws.cell(row=r, column=col).font = font_bold
                    
    elif report_type == "st3":
        ws.title = "СТ-3 Өмч"
        ws.append([company_name])
        ws.append(["СТ-3: Өмчийн өөрчлөлтийн тайлан"])
        ws.append([f"Тайлант хугацаа: {date_from_str} - {date_to_str}"])
        ws.append(["Хэмжих нэгж: Төгрөг (₮)"])
        ws.append([])
        
        ws.cell(row=1, column=1).font = font_title
        ws.cell(row=2, column=1).font = font_bold
        
        headers = ["Үзүүлэлт", "Хувь нийлүүлсэн хөрөнгө", "Хуримтлагдсан ашиг", "Нийт өмч"]
        ws.append(headers)
        for col_idx in range(1, 5):
            cell = ws.cell(row=6, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            
        eq = reports.statement_of_changes_in_equity(session, company_id, d_from, d_to)
        curr = eq["current"]
        
        rows = [
            ("Эхний үлдэгдэл", curr["start"]),
            ("Тайлант үеийн цэвэр ашиг", curr["net_income"]),
            ("Бусад өөрчлөлт (Ногдол ашиг г.м)", curr["changes"]),
            ("Эцсийн үлдэгдэл", curr["end"]),
        ]
        for name, data in rows:
            ws.append([name, _minor_to_unit(data["capital"]), _minor_to_unit(data["retained"]), _minor_to_unit(data["total"])])
            r = ws.max_row
            ws.cell(row=r, column=1).alignment = align_left
            ws.cell(row=r, column=2).alignment = align_right
            ws.cell(row=r, column=3).alignment = align_right
            ws.cell(row=r, column=4).alignment = align_right
            ws.cell(row=r, column=2).number_format = '#,##0.00'
            ws.cell(row=r, column=3).number_format = '#,##0.00'
            ws.cell(row=r, column=4).number_format = '#,##0.00'
            if "Эцсийн" in name or "Нийт" in name:
                for col in range(1, 5):
                    ws.cell(row=r, column=col).font = font_bold
                    
    elif report_type == "st4":
        ws.title = "СТ-4 Мөнгөн гүйлгээ"
        ws.append([company_name])
        ws.append(["СТ-4: Мөнгөн гүйлгээний тайлан (Шууд арга)"])
        ws.append([f"Тайлант хугацаа: {date_from_str} - {date_to_str}"])
        ws.append(["Хэмжих нэгж: Төгрөг (₮)"])
        ws.append([])
        
        ws.cell(row=1, column=1).font = font_title
        ws.cell(row=2, column=1).font = font_bold
        
        headers = ["Үзүүлэлт", "Тайлант үеийн дүн"]
        ws.append(headers)
        for col_idx in range(1, 3):
            cell = ws.cell(row=6, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            
        cf = reports.cash_flow_statement(session, company_id, d_from, d_to)
        curr = cf["current"]
        
        rows = [
            ("1. Үйл ажиллагааны мөнгөн гүйлгээ", None),
            ("  Борлуулалт, үйлчилгээний орлогоос орсон мөнгө", curr["operating_inflow"]),
            ("  Материал, цалин, даатгал, татварын зарлагад гарсан мөнгө", -curr["operating_outflow"]),
            ("Үйл ажиллагааны цэвэр мөнгөн гүйлгээ", curr["operating_net"]),
            ("2. Хөрөнгө оруулалтын мөнгөн гүйлгээ", None),
            ("  Үндсэн хөрөнгө борлуулснаас орсон мөнгө", curr["investing_inflow"]),
            ("  Үндсэн хөрөнгө худалдан авахад гарсан мөнгө", -curr["investing_outflow"]),
            ("Хөрөнгө оруулалтын цэвэр мөнгөн гүйлгээ", curr["investing_net"]),
            ("3. Санхүүжилтийн мөнгөн гүйлгээ", None),
            ("  Капитал нэмэгдүүлсэн, зээл авснаас орсон мөнгө", curr["financing_inflow"]),
            ("  Ногдол ашиг олгосон, зээл эргэн төлөхөд гарсан мөнгө", -curr["financing_outflow"]),
            ("Санхүүжилтийн цэвэр мөнгөн гүйлгээ", curr["financing_net"]),
            ("Мөнгө, түүнтэй адилтгах хөрөнгийн цэвэр өөрчлөлт", curr["net_change"]),
        ]
        for name, val in rows:
            if val is None:
                ws.append([name, ""])
                r = ws.max_row
                ws.cell(row=r, column=1).font = font_bold
            else:
                ws.append([name, _minor_to_unit(val)])
                r = ws.max_row
                ws.cell(row=r, column=1).alignment = align_left
                ws.cell(row=r, column=2).alignment = align_right
                ws.cell(row=r, column=2).number_format = '#,##0.00'
                if "цэвэр" in name.lower() or "өөрчлөлт" in name.lower() or "Үйл ажиллагааны цэвэр" in name:
                    for col in range(1, 3):
                        ws.cell(row=r, column=col).font = font_bold
                        
    elif report_type == "stock":
        ws.title = "Бараа материалын үлдэгдэл"
        ws.append([company_name])
        ws.append(["Бараа материалын үлдэгдлийн тайлан"])
        ws.append([f"Тайлант хугацаа: {date_to_str} байдлаар"])
        ws.append(["Хэмжих нэгж: Төгрөг (₮)"])
        ws.append([])
        
        ws.cell(row=1, column=1).font = font_title
        ws.cell(row=2, column=1).font = font_bold
        
        headers = ["Барааны код", "Барааны нэр", "Үлдэгдэл тоо", "Дундаж өртөг", "Нийт өртөг үлдэгдэл"]
        ws.append(headers)
        for col_idx in range(1, 6):
            cell = ws.cell(row=6, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            
        from . import inventory
        stock = inventory.get_stock_balances(session, company_id)
        for i in stock:
            ws.append([
                i["code"], 
                i["name"], 
                i["qty"], 
                _minor_to_unit(i["avg_cost_minor"]), 
                _minor_to_unit(i["total_cost_minor"])
            ])
            r = ws.max_row
            ws.cell(row=r, column=1).alignment = align_left
            ws.cell(row=r, column=2).alignment = align_left
            ws.cell(row=r, column=3).alignment = align_right
            ws.cell(row=r, column=4).alignment = align_right
            ws.cell(row=r, column=5).alignment = align_right
            ws.cell(row=r, column=4).number_format = '#,##0.00'
            ws.cell(row=r, column=5).number_format = '#,##0.00'
            
    elif report_type == "aging":
        ws.title = "Авлагын насжилт"
        ws.append([company_name])
        ws.append(["Авлагын насжилтын тайлан"])
        ws.append([f"Тайлант хугацаа: {date_to_str} байдлаар"])
        ws.append(["Хэмжих нэгж: Төгрөг (₮)"])
        ws.append([])
        
        ws.cell(row=1, column=1).font = font_title
        ws.cell(row=2, column=1).font = font_bold
        
        headers = ["Харилцагч", "Нэхэмжлэх", "Хэтэрсэн хоног", "0-30", "31-60", "61-90", "91-120", "120+"]
        ws.append(headers)
        for col_idx in range(1, 9):
            cell = ws.cell(row=6, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            
        from . import partners
        aging = partners.get_aging_report(session, company_id, kind="sales")
        for a in aging:
            ws.append([
                a["counterparty"],
                a["number"],
                a["overdue_days"],
                _minor_to_unit(a["buckets"][0]) if a["buckets"][0] else 0.0,
                _minor_to_unit(a["buckets"][1]) if a["buckets"][1] else 0.0,
                _minor_to_unit(a["buckets"][2]) if a["buckets"][2] else 0.0,
                _minor_to_unit(a["buckets"][3]) if a["buckets"][3] else 0.0,
                _minor_to_unit(a["buckets"][4]) if a["buckets"][4] else 0.0,
            ])
            r = ws.max_row
            ws.cell(row=r, column=1).alignment = align_left
            ws.cell(row=r, column=2).alignment = align_left
            ws.cell(row=r, column=3).alignment = align_right
            for c in range(4, 9):
                ws.cell(row=r, column=c).alignment = align_right
                ws.cell(row=r, column=c).number_format = '#,##0.00'
                
    elif report_type == "vat":
        ws.title = "НӨАТ тайлан"
        ws.append([company_name])
        ws.append(["НӨАТ-ын тайлан (ТТ-03а)"])
        ws.append([f"Тайлант хугацаа: {date_from_str.split('-')[0]} он"])
        ws.append(["Хэмжих нэгж: Төгрөг (₮)"])
        ws.append([])
        
        ws.cell(row=1, column=1).font = font_title
        ws.cell(row=2, column=1).font = font_bold
        
        headers = ["Үзүүлэлт", "Дүн"]
        ws.append(headers)
        for col_idx in range(1, 3):
            cell = ws.cell(row=6, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            
        from . import vat
        v = vat.calculate_tt03a(session, company_id, int(date_from_str.split('-')[0]))
        rows = [
            ("Мөр 31 — Ногдуулсан НӨАТ", v["rows"]["31_nogduulsan_niit"]),
            ("Мөр 49 — Хасагдах НӨАТ", v["rows"]["49_hasagdah_noat"]),
            ("Мөр 64 — Төлбөл зохих", v["rows"]["64_etssiin_tolboh"]),
            ("Мөр 65 — Буцаан авах", v["rows"]["65_etssiin_butsaan_avah"]),
        ]
        for name, val in rows:
            ws.append([name, _minor_to_unit(val)])
            r = ws.max_row
            ws.cell(row=r, column=1).alignment = align_left
            ws.cell(row=r, column=2).alignment = align_right
            ws.cell(row=r, column=2).number_format = '#,##0.00'
            if "Төлбөл зохих" in name or "Буцаан авах" in name:
                for col in range(1, 3):
                    ws.cell(row=r, column=col).font = font_bold

    elif report_type in ("notes", "nbfi", "tt02", "tt11"):
        if report_type == "notes":
            ws.title = "Тодруулга"
            ws.append([company_name, "Санхүүгийн тайлангийн 12 Тодруулга"])
            fn = reports.financial_notes(session, company_id, d_from, d_to)
            ws.append([])
            ws.append(["Тодруулга 1: Мөнгө ба мөнгөн адилтгах хөрөнгө", _minor_to_unit(fn["note_1_cash"]["total_minor"])])
            ws.append(["Тодруулга 2: Авлага", _minor_to_unit(fn["note_2_receivables"]["total_minor"])])
            ws.append(["Тодруулга 3: Бараа материал", _minor_to_unit(fn["note_3_inventory"]["total_minor"])])
            ws.append(["Тодруулга 4: Үндсэн хөрөнгө (Цэвэр үлдэгдэл)", _minor_to_unit(fn["note_4_ppe"]["net_book_value_minor"])])
            ws.append(["Тодруулга 5: Өр төлбөр", _minor_to_unit(fn["note_5_liabilities"]["total_minor"])])
            ws.append(["Тодруулга 6: Зардал (Борлуулалт ба Ерөнхий удирдлага)", _minor_to_unit(fn["note_6_expenses"]["total_minor"])])
        elif report_type == "nbfi":
            ws.title = "СЗХ ББСБ Тайлан"
            ws.append([company_name, "ББСБ-ын СЗХ-ны тусгай тайлан ба Мэдэгдэл"])
            nbfi = reports.nbfi_financial_reports(session, company_id, d_to)
            ws.append([])
            ws.append(["СТ-1 ББСБ Нийт актив", _minor_to_unit(nbfi["st1_nbfi_balance_sheet"]["total_assets_minor"])])
            ws.append(["СТ-2 ББСБ Цэвэр ашиг", _minor_to_unit(nbfi["st2_nbfi_income_statement"]["net_income_minor"])])
            ws.append(["Мэдэгдэл", nbfi["representation_statement"]["statement_title"]])
        elif report_type == "tt02":
            ws.title = "ААНОАТ ТТ-02"
            t2 = reports.aanoat_tt02_report(session, company_id, int(date_from_str.split('-')[0]) if date_from_str else 2026)
            ws.append([company_name, t2["form_name"]])
            ws.append([])
            ws.append(["Нийт орлого", _minor_to_unit(t2["gross_revenue_minor"])])
            ws.append(["Зөвшөөрөгдөх хасагдах зардал", _minor_to_unit(t2["allowable_expenses_minor"])])
            ws.append(["Татвар ногдуулах орлого", _minor_to_unit(t2["taxable_income_minor"])])
            ws.append(["Ногдуулсан ААНОАТ (10%)", _minor_to_unit(t2["calculated_tax_minor"])])
        elif report_type == "tt11":
            ws.title = "ХХОАТ ТТ-11"
            t11 = reports.haoat_tt11_report(session, company_id, int(date_from_str.split('-')[0]) if date_from_str else 2026)
            ws.append([company_name, t11["form_name"]])
            ws.append([])
            ws.append(["Ажилтны тоо", t11.get("employee_count", 0)])
            ws.append(["Олгосон нийт цалин", _minor_to_unit(t11["gross_salary_minor"])])
            ws.append(["Суутгасан НДШ", _minor_to_unit(t11["ndsh_deduction_minor"])])
            ws.append(["Татвар ногдуулах орлого", _minor_to_unit(t11["taxable_income_minor"])])
            ws.append(["Суутгасан ХХОАТ", _minor_to_unit(t11["haoat_tax_minor"])])
            ws.append(["Гарт олгосон", _minor_to_unit(t11.get("net_pay_minor", 0))])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_path = Path(tmp.name)
    tmp.close()
    wb.save(tmp_path)
    return tmp_path

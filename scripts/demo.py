# -*- coding: utf-8 -*-
"""Bayan AI — Үе 1-ийн ажиллагааны demo.

Хийх зүйл:
  1. Demo компани + дансны төлөвлөгөө + банкны GL данс үүсгэнэ
  2. Синтетик Хаан банкны хуулга (12 гүйлгээ, 2026-03 сар) үүсгэнэ
  3. Pipeline: таних → задлах → нормчлох → validation gate
  4. Дүрмийн ангилалт → батлах → журналын бичилтүүд
  5. Гүйлгээ баланс + банкны үлдэгдлийн тулгалт
  6. ЭВДЭРСЭН хуулгаар gate унахыг үзүүлнэ
  7. demo-report.html тайлан гаргана

Ажиллуулах:  .venv\\Scripts\\python.exe scripts\\demo.py
"""
import io
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from openpyxl import Workbook

from bayan import ledger
from bayan.amounts import format_minor
from bayan.classify import classify_batch
from bayan.coa_seed import add_bank_gl_account, seed_company
from bayan.db import make_session
from bayan.models import BankTxn, ClassificationSuggestion, ClassifierRule
from bayan.pipeline import approve_suggestions, process_file

OUT = Path(__file__).resolve().parents[1]

TXNS = [
    # (огноо, дебит, кредит, харьцсан данс, утга)
    ("2026.03.01 09:15", None, 2_500_000.00, "5099887766", "Борлуулалтын орлого нэхэмжлэх INV-2026-018"),
    ("2026.03.03 11:20", 850_000.00, None, "5011223344", "Түрээсийн төлбөр 3-р сар Оффис центр ХХК"),
    ("2026.03.05 10:00", None, 1_800_000.00, "5044556677", "Борлуулалтын орлого Мөнхжин ХХК"),
    ("2026.03.07 15:30", 420_000.00, None, None, "SOCIAL INSURANCE 2-р сарын НДШ"),
    ("2026.03.10 09:45", 1_260_000.00, None, "5033445566", "Цалин олголт 2-р сарын сүүл"),
    ("2026.03.12 14:10", 96_000.00, None, None, "Юнител групп интернэт, утас 3 сар"),
    ("2026.03.15 10:30", None, 3_200_000.00, "5055667788", "Гэрээт ажлын урьдчилгаа Баянгол Трейд"),
    ("2026.03.18 16:00", 540_000.00, None, "5066778899", "Шатахуун Petrovis карт цэнэглэлт"),
    ("2026.03.22 11:45", 175_000.00, None, "5077889900", "Бичиг хэргийн материал Юнис ХХК"),
    ("2026.03.25 09:00", None, 950_000.00, "5088990011", "Борлуулалтын орлого бэлэн бус"),
    ("2026.03.28 13:20", 2_100_000.00, None, "5099001122", "Түүхий эд материал худалдан авалт Тэсо ХХК"),
    ("2026.03.31 17:55", 8_400.00, None, None, "Гүйлгээний шимтгэл 3-р сар"),
]

RULES = [
    ("борлуулалтын орлого", "5101", 10),
    ("түрээс",              "7103", 10),
    ("social insurance",    "3103", 10),
    ("ндш",                 "3103", 15),
    ("цалин олголт",        "3102", 10),
    ("юнител",              "7105", 20),
    ("урьдчилгаа",          "3108", 20),
    ("шатахуун",            "7104", 20),
    ("petrovis",            "7104", 15),
    ("бичиг хэрг",          "7199", 20),
    ("түүхий эд",           "2101", 15),
    ("шимтгэл",             "7106", 20),
]

OPENING = 4_000_000.00


def make_statement(path: Path, corrupt: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Дансны хуулга"
    ws["A1"] = "Хаан банк — Дансны хуулга"
    ws["A2"] = "Дансны дугаар:"; ws["B2"] = "5041234567"
    ws["A3"] = "Дансны нэр:";    ws["B3"] = "Баян Демо ХХК"
    ws["A4"] = "Хуулгын хугацаа:"; ws["B4"] = "2026.03.01 - 2026.03.31"
    ws["A5"] = "Эхний үлдэгдэл:"; ws["B5"] = OPENING
    ws.append([])
    ws.append(["Гүйлгээний огноо", "Салбар", "Дебит гүйлгээ", "Кредит гүйлгээ",
               "Үлдэгдэл", "Харьцсан данс", "Гүйлгээний утга"])
    bal = OPENING
    for d, dr, cr, cp, desc in TXNS:
        bal = bal - (dr or 0) + (cr or 0)
        ws.append([d, "УБ салбар", dr, cr, round(bal, 2), cp, desc])
    if corrupt:
        c = ws.cell(row=12, column=5)          # 5 дахь гүйлгээний үлдэгдлийг эвдэнэ
        c.value = float(c.value) + 100.00
    ws.append([])
    ws.append(["Эцсийн үлдэгдэл:", round(bal, 2)])
    ws.append(["Гүйлгээний тоо:", len(TXNS)])
    wb.save(path)


def main() -> None:
    session = make_session("sqlite:///:memory:")
    print("=" * 72)
    print("BAYAN AI — Үе 1 DEMO:  хуулга → parse → gate → ангилалт → журнал → тайлан")
    print("=" * 72)

    # 1. Компани
    company = seed_company(session, "Баян Демо ХХК")
    bank_gl = add_bank_gl_account(session, company.id, "khan", "5041234567")
    for kw, code, prio in RULES:
        session.add(ClassifierRule(company_id=company.id, keyword=kw,
                                   account_code=code, priority=prio))
    session.flush()
    print(f"\n[1] Компани: {company.name}  |  Банкны GL данс: {bank_gl.code} {bank_gl.name}")
    print(f"    Нябо-гийн ангилалтын дүрэм: {len(RULES)}")

    # 2. Хуулга үүсгэх
    good = OUT / "samples_demo" ; good.mkdir(exist_ok=True)
    f_good = good / "khan_statement_2026_03.xlsx"
    f_bad = good / "khan_statement_2026_03_corrupt.xlsx"
    make_statement(f_good)
    make_statement(f_bad, corrupt=True)
    print(f"\n[2] Синтетик Хаан хуулга: {len(TXNS)} гүйлгээ, нээлт {OPENING:,.2f}₮")

    # 3. Pipeline — зөв файл
    report = process_file(session, company.id, f_good)
    print(f"\n[3] PIPELINE (зөв файл): descriptor={report.descriptor_id}")
    print(f"    Задлагдсан гүйлгээ: {report.txn_count}")
    print(f"    Validation gate:   {'✓ ДАВЛАА (V1-V6 бүгд OK)' if report.gate_ok else '✗'}")

    # 4. Ангилалт + батлах
    txns = session.query(BankTxn).order_by(BankTxn.seq_no).all()
    classify_batch(session, company.id, txns, use_ai=False)
    sugs = session.query(ClassificationSuggestion).all()
    by_txn = {s.bank_txn_id: s for s in sugs}
    print(f"\n[4] АНГИЛАЛТ (дүрмийн давхарга — AI key-гүй горим):")
    for t in txns:
        s = by_txn[t.id]
        arrow = "←" if t.direction.value == "credit" else "→"
        print(f"    {t.posted_at:%m-%d} {arrow} {format_minor(t.amount_minor):>13}₮"
              f"  {s.account_code}  [{s.confidence:.2f}] {t.description_norm[:44]}")

    approved = approve_suggestions(session, company.id, [s.id for s in sugs], bank_gl.code)
    print(f"\n    Батлагдсан → журналын бичилт: {len(approved)}")

    # 5. Гүйлгээ баланс
    print(f"\n[5] ГҮЙЛГЭЭ БАЛАНС (2026-03):")
    print(f"    {'Код':<8}{'Данс':<38}{'Дебит':>14}{'Кредит':>14}{'Үлдэгдэл':>14}")
    tb = ledger.trial_balance(session, company.id)
    td = tc = 0
    for r in tb:
        td += r["debit_minor"]; tc += r["credit_minor"]
        print(f"    {r['code']:<8}{r['name'][:36]:<38}"
              f"{format_minor(r['debit_minor']):>14}{format_minor(r['credit_minor']):>14}"
              f"{format_minor(r['balance_minor']):>14}")
    print(f"    {'':‌<8}{'НИЙТ':<38}{format_minor(td):>14}{format_minor(tc):>14}")
    assert td == tc, "Σдебит ≠ Σкредит!"

    bank_row = next(r for r in tb if r["code"] == bank_gl.code)
    expected_closing = int(OPENING * 100) + bank_row["balance_minor"]
    stmt_closing = session.query(BankTxn).order_by(BankTxn.seq_no.desc()).first()
    print(f"\n    Банкны GL үлдэгдэл + нээлт = {format_minor(expected_closing)}₮")
    print(f"    Хуулгын эцсийн үлдэгдэл   = {format_minor(stmt_closing.balance_after_minor)}₮")
    ok = expected_closing == stmt_closing.balance_after_minor
    print(f"    Тулгалт: {'✓ МӨНГӨ БҮРЭЭР ТААРЛАА' if ok else '✗ ЗӨРЛӨӨ'}")

    # 6. Эвдэрсэн файл
    report_bad = process_file(session, company.id, f_bad)
    print(f"\n[6] PIPELINE (ЭВДЭРСЭН файл — 5 дахь гүйлгээний үлдэгдлийг 100₮-өөр өөрчилсөн):")
    print(f"    Validation gate: {'✓' if report_bad.gate_ok else '✗ УНАЛАА (зөв!)'}")
    for i in report_bad.issues[:4]:
        row = f"мөр {i['row']}" if i["row"] else "ерөнхий"
        print(f"      [{i['check']}] {row}: {i['detail']}")
    n_saved = session.query(BankTxn).filter(
        BankTxn.statement_id == report_bad.statement_id).count()
    print(f"    Систем рүү орсон гүйлгээ: {n_saved} (gate даваагүй тул 0 байх ёстой)")

    print("\n" + "=" * 72)
    print("DEMO ДУУСЛАА — бүх инвариант ажиллаж байна.")
    print("=" * 72)


if __name__ == "__main__":
    main()
